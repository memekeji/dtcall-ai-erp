# 标准库导入
import json
import logging
import uuid
from datetime import datetime, timedelta

# Django核心导入
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin
from django.core.exceptions import ValidationError
from django.db import models, transaction
from django.db.models import Prefetch, Exists, OuterRef, Q, Count
from django.http import JsonResponse, HttpResponseRedirect
from django.shortcuts import render, redirect, get_object_or_404
from django.urls import reverse_lazy, reverse
from django.utils import timezone
from django.utils.decorators import method_decorator
from django.views.decorators.cache import never_cache
from django.views.decorators.csrf import csrf_exempt
from django.views.decorators.http import require_POST
from django.views.generic import (
    View, ListView, CreateView, DetailView, 
    UpdateView, DeleteView, TemplateView
)

# 系统日志导入
from apps.user.models import SystemLog

# 本地应用导入
from .models import (
    Customer, CustomerGrade, CustomerSource, CustomerIntent, SpiderTask, 
    Contact, FollowRecord, CustomerField, CustomerCustomFieldValue,
    CustomerOrder, CustomerContract, CustomerInvoice, CustomerOrderCustomFieldValue,
    FollowField, OrderField
)
# 财务模块导入
from apps.finance.models import Income, Payment
# 用户模型导入
from apps.user.models.admin import Admin
from .forms import CustomerForm, ContactFormSet, CustomerFieldForm, FollowFieldForm, OrderFieldForm
from .forms import CustomerSourceForm, CustomerGradeForm, CustomerIntentForm
from .serializers import CustomerFieldSerializer
from .filters import CustomerFilter

logger = logging.getLogger(__name__)


class CustomerListView(LoginRequiredMixin, ListView):
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = Customer
    template_name = 'customer/customer_list.html'
    context_object_name = 'customers'
    
    def get_queryset(self):
        # 只返回未删除的客户记录
        queryset = super().get_queryset()
        queryset = queryset.filter(delete_time=0)
        
        # 添加数据权限过滤
        user = self.request.user
        
        # 超级管理员可以查看所有客户
        if hasattr(user, 'is_superuser') and user.is_superuser:
            return queryset
        
        # 数据权限过滤：只能查看自己的客户及共享给自己的客户
        queryset = queryset.filter(
            models.Q(belong_uid=user.id) | 
            models.Q(share_ids__contains=str(user.id))
        )
        
        return queryset
        
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # 获取客户来源和等级信息
        context['sources'] = CustomerSource.objects.filter(status=1, delete_time=0)
        context['grades'] = CustomerGrade.objects.filter(status=1, delete_time=0)
        
        # 获取客户意向信息
        context['intents'] = CustomerIntent.objects.filter(status=1, delete_time=0)
        
        # 从JSON文件读取完整的省市数据
        import json
        import os
        
        json_file_path = os.path.join(settings.BASE_DIR, 'static', 'json', '全国省市区.json')
        
        try:
            with open(json_file_path, 'r', encoding='utf-8') as f:
                province_city_data = json.load(f)
            
            # 获取所有省份数据
            provinces_data = province_city_data.get('00', {})
            context['provinces'] = list(provinces_data.values())
            
            # 构建省市映射关系
            province_city_map = {}
            for province_key in province_city_data:
                if province_key != '00':  # 跳过省份数据
                    # 获取省份名称（通过键值对应）
                    province_name = provinces_data.get(province_key, '')
                    if province_name:
                        cities_data = province_city_data.get(province_key, {})
                        province_city_map[province_name] = list(cities_data.values())
            
            context['cities'] = province_city_map
            
        except Exception as e:
            # 如果JSON文件读取失败，使用客户数据中的省市数据作为备选
            provinces = Customer.objects.filter(delete_time=0, province__isnull=False).exclude(province='').values_list('province', flat=True).distinct()
            
            # 构建简单的省市映射（基于现有客户数据）
            province_city_map = {}
            for province in provinces:
                cities = Customer.objects.filter(delete_time=0, province=province, city__isnull=False).exclude(city='').values_list('city', flat=True).distinct()
                province_city_map[province] = sorted(cities)
            
            context['provinces'] = sorted(provinces)
            context['cities'] = province_city_map
        
        # 获取启用的客户字段，用于动态生成表格列
        custom_fields = CustomerField.objects.filter(
            status=True, 
            delete_time=0, 
            is_list_display=True
        ).order_by('sort', 'id')
        context['custom_fields'] = custom_fields
        
        # 添加用户权限信息
        context['is_superuser'] = hasattr(self.request.user, 'is_superuser') and self.request.user.is_superuser
        
        return context


class CustomerListSimpleView(LoginRequiredMixin, TemplateView):
    login_url = '/user/login/'
    redirect_field_name = 'next'
    template_name = 'customer/customer_list_simple.html'


class CustomerListDebugView(LoginRequiredMixin, TemplateView):
    login_url = '/user/login/'
    redirect_field_name = 'next'
    template_name = 'customer/customer_list_debug.html'

    def get_queryset(self):
        queryset = super().get_queryset()
        queryset = queryset.filter(delete_time=0)
        queryset = queryset.prefetch_related(
            Prefetch('custom_fields',
                queryset=CustomerCustomFieldValue.objects.select_related('field'),
                to_attr='custom_field_values'
            )
        )
        return queryset

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['sources'] = CustomerSource.objects.filter(status=1, delete_time=0)
        context['grades'] = CustomerGrade.objects.filter(status=1, delete_time=0)
        
        # 获取启用的客户字段，用于动态生成表格列
        custom_fields = CustomerField.objects.filter(
            status=True, 
            delete_time=0, 
            is_list_display=True
        ).order_by('sort', 'id')
        context['custom_fields'] = custom_fields
        
        return context


class CustomerListDataView(LoginRequiredMixin, View):
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def get(self, request):
        try:
            # 获取客户列表数据，并使用annotate添加实体关联计数
            queryset = Customer.objects.filter(delete_time=0)
            
            # 添加实体关联计数
            queryset = queryset.annotate(
                # 订单计数
                order_count=Count('orders', filter=models.Q(orders__delete_time=0)),
                # 合同计数
                contract_count=Count('contracts', filter=models.Q(contracts__delete_time=0)),
                # 项目计数
                project_count=Count('projects'),
                # 发票计数
                invoice_count=Count('invoices', filter=models.Q(invoices__delete_time=0))
            )
            
            # 添加数据权限过滤
            user = self.request.user
            
            # 数据权限过滤：
            # 1. 超级管理员：可以查看所有客户，但不包括已移入公海的客户（belong_uid=0）
            # 2. 普通用户：只能查看自己的客户及共享给自己的客户
            if hasattr(user, 'is_superuser') and user.is_superuser:
                # 超级管理员：排除已移入公海的客户
                queryset = queryset.filter(belong_uid__gt=0)
            else:
                # 普通用户：只能查看自己的客户及共享给自己的客户
                queryset = queryset.filter(
                    models.Q(belong_uid=user.id) | 
                    models.Q(share_ids__contains=str(user.id))
                )
            
            # 获取视图类型（列表视图或卡片视图）
            view_type = request.GET.get('view_type', 'list')  # 默认为列表视图
            
            # 根据视图类型设置排序规则
            if view_type == 'card':
                # 卡片视图：按ID降序排序，因为services__sort无法解析
                queryset = queryset.order_by('-id')
            else:
                # 列表视图：按ID降序排序
                queryset = queryset.order_by('-id')
            
            # 获取自定义字段定义
            custom_fields = CustomerField.objects.filter(delete_time=0, status=True)
            serialized_fields = CustomerFieldSerializer(custom_fields, many=True).data
            
            # 处理搜索条件
            customer_name = request.GET.get('customer_name', '')
            contact_name = request.GET.get('contact_name', '')
            phone = request.GET.get('phone', '')
            
            if customer_name:
                queryset = queryset.filter(name__icontains=customer_name)
            if contact_name:
                queryset = queryset.filter(contacts__contact_person__icontains=contact_name)
            if phone:
                queryset = queryset.filter(contacts__phone__icontains=phone)
            
            # 处理筛选条件
            customer_source = request.GET.get('customer_source', '')
            customer_grade = request.GET.get('customer_grade', '')
            customer_intent = request.GET.get('customer_intent', '')
            province = request.GET.get('province', '')
            city = request.GET.get('city', '')
            
            if customer_source:
                queryset = queryset.filter(customer_source__title=customer_source)
            if customer_grade:
                queryset = queryset.filter(grade_id__in=CustomerGrade.objects.filter(title=customer_grade).values_list('id', flat=True))
            if customer_intent:
                queryset = queryset.filter(services_id__in=CustomerIntent.objects.filter(name=customer_intent).values_list('id', flat=True))
            if province:
                queryset = queryset.filter(province=province)
            if city:
                queryset = queryset.filter(city=city)
            
            # 处理自定义字段筛选
            for field in custom_fields:
                filter_key = f'custom_filter_{field.id}'
                filter_value = request.GET.get(filter_key, '')
                if filter_value:
                    queryset = queryset.filter(custom_fields__field_id=field.id, custom_fields__value=filter_value)
            
            # 添加预取操作
            queryset = queryset.prefetch_related(
                'contacts',
                Prefetch('custom_fields',
                    queryset=CustomerCustomFieldValue.objects.select_related('field'),
                    to_attr='custom_field_values'
                ),
                Prefetch('follow_records',
                    queryset=FollowRecord.objects.filter(delete_time=0).order_by('-follow_time'),
                    to_attr='latest_follow_records'
                )
            )

            # 处理分页
            page = int(request.GET.get('page', 1))
            limit = int(request.GET.get('limit', 10))
            start = (page - 1) * limit
            end = start + limit

            total_count = queryset.count()
            paginated_queryset = queryset[start:end]

            # 获取基础数据映射
            sources_map = {s.id: s.title for s in CustomerSource.objects.filter(delete_time=0)}
            grades_map = {g.id: g.title for g in CustomerGrade.objects.filter(delete_time=0)}
            intents_map = {i.id: i.name for i in CustomerIntent.objects.filter(delete_time=0)}
            
            # 构建用户ID到姓名的映射
            user_ids = list(queryset.values_list('belong_uid', flat=True).distinct())
            users_map = {u.id: u.name for u in Admin.objects.filter(id__in=user_ids)}
            
            # 格式化数据
            items = []
            for item in paginated_queryset:
                # 获取主要联系人信息
                primary_contact = item.contacts.filter(is_primary=True).first()
                if not primary_contact:
                    primary_contact = item.contacts.first()
                
                contact_name = primary_contact.contact_person if primary_contact else ''
                phone = primary_contact.phone if primary_contact else ''
                email = primary_contact.email if primary_contact else ''
                
                # 获取最近跟进时间
                latest_followup_time = ''
                if hasattr(item, 'latest_follow_records') and item.latest_follow_records:
                    latest_follow_record = item.latest_follow_records[0]  # 按时间降序排列，第一个就是最新的
                    latest_followup_time = latest_follow_record.follow_time.strftime('%Y-%m-%d %H:%M:%S') if latest_follow_record.follow_time else ''
                
                # 获取客户归属信息
                customer_owner = ''
                if item.belong_uid and item.belong_uid > 0:
                    customer_owner = users_map.get(item.belong_uid, '')
                else:
                    customer_owner = '公海客户'
                
                item_data = {
                'id': item.id,
                'name': item.name,
                'contact_name': contact_name,
                'phone': phone,
                'email': email,
                'address': item.address,
                'create_time': item.create_time.strftime('%Y-%m-%d %H:%M:%S') if item.create_time else '',
                'customer_source': sources_map.get(item.customer_source_id, ''),
                'customer_grade': grades_map.get(item.grade_id, ''),
                'customer_intent': intents_map.get(item.services_id, ''),
                'customer_owner': customer_owner,  # 添加客户归属字段
                'latest_followup_time': latest_followup_time,
                'intent_sort': getattr(item, 'intent_sort', 999),  # 添加意向排序值
                # 添加实体关联计数
                'order_count': getattr(item, 'order_count', 0),
                'contract_count': getattr(item, 'contract_count', 0),
                'project_count': getattr(item, 'project_count', 0),
                'invoice_count': getattr(item, 'invoice_count', 0),
                'payment_count': getattr(item, 'payment_count', 0)
            }
                
                # 添加自定义字段值
                for cfv in item.custom_field_values:
                    item_data[f'custom_{cfv.field.id}'] = cfv.value
                
                items.append(item_data)

            return JsonResponse({
                'code': 0,
                'msg': '',
                'count': total_count,
                'data': items,
                'custom_fields': serialized_fields
            })
        except Exception as e:
            logger.error(f"Error processing customer list data: {str(e)}", exc_info=True)
            return JsonResponse({
                'code': 500,
                'msg': f'Server error: {str(e)}',
                'data': []
            }, status=500)


class CustomerCreateView(LoginRequiredMixin, CreateView):
    form_class = CustomerForm
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = Customer
    template_name = 'customer/customer_form.html'
    success_url = reverse_lazy('customer:customer_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # 获取启用的自定义字段
        custom_fields = CustomerField.objects.filter(status=True, delete_time=0).order_by('sort')
        
        # 为每个字段添加选项列表和当前值
        for field in custom_fields:
            if field.options:
                field.options_list = [opt.strip() for opt in field.options.split('\n') if opt.strip()]
            else:
                field.options_list = []
            field.current_value = ''
        
        context['custom_fields'] = custom_fields
        
        # 联系人表单集
        if self.request.POST:
            context['contact_formset'] = ContactFormSet(self.request.POST, instance=self.object)
        else:
            context['contact_formset'] = ContactFormSet(instance=self.object)
        
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        contact_formset = context['contact_formset']
        
        with transaction.atomic():
            # 设置客户归属用户ID为当前登录用户
            form.instance.belong_uid = self.request.user.id
            form.instance.admin_id = self.request.user.id
            # 设置初始状态值确保通过列表过滤条件
            form.instance.delete_time = 0
            form.instance.intent_status = 1
            # 设置归属时间为当前时间
            form.instance.belong_time = int(timezone.now().timestamp())
            
            self.object = form.save()
            
            # 处理主要联系人的单选逻辑
            primary_contact_value = self.request.POST.get('primary_contact')
            
            # 保存联系人信息
            if contact_formset.is_valid():
                contact_formset.instance = self.object
                contacts = contact_formset.save()
                
                # 处理主要联系人设置
                if primary_contact_value and contacts:
                    # 先将所有联系人设为非主要
                    Contact.objects.filter(customer=self.object).update(is_primary=False)
                    
                    # 根据primary_contact的值设置主要联系人
                    if primary_contact_value.isdigit():
                        # 如果是数字，表示是新添加的联系人索引
                        contact_index = int(primary_contact_value)
                        if contact_index < len(contacts):
                            contacts[contact_index].is_primary = True
                            contacts[contact_index].save()
                    else:
                        # 如果不是数字，可能是联系人ID
                        try:
                            contact_id = int(primary_contact_value)
                            Contact.objects.filter(id=contact_id, customer=self.object).update(is_primary=True)
                        except (ValueError, TypeError):
                            pass
            else:
                return self.form_invalid(form)
            
            # 处理自定义字段
            self.save_custom_fields()
            
            # 需求案例1：创建客户后自动生成待签约合同、待确认订单、待收款、待开票、项目
            try:
                self._auto_generate_related_records(self.object, self.request.user)
            except Exception as auto_gen_error:
                # 记录错误但不影响客户创建
                logger.error(f"自动生成相关记录失败: {str(auto_gen_error)}")
            
            # 添加操作日志
            SystemLog.objects.create(
                user=self.request.user,
                log_type='create',
                module='客户管理',
                action='创建客户',
                content=f'成功创建客户: {self.object.name}',
                ip_address=self.request.META.get('REMOTE_ADDR'),
                user_agent=self.request.META.get('HTTP_USER_AGENT')
            )
        
        messages.success(self.request, '客户添加成功！')
        # 直接使用reverse生成URL进行重定向，确保重定向正确执行
        from django.urls import reverse
        return HttpResponseRedirect(reverse('customer:customer_list'))
    
    def _auto_generate_related_records(self, customer, user):
        """
        自动生成与客户相关的记录：待签约合同、待确认订单、待收款、待开票、项目
        """
        import time
        
        # 1. 生成待签约合同记录（CustomerContract）
        try:
            from apps.customer.models import CustomerContract
            # 创建待签约合同记录
            contract_record = CustomerContract.objects.create(
                customer_id=customer.id,
                name=f"{customer.name}合同",
                contract_number=f"CONT-CUST-{customer.id}-{int(time.time())}",
                amount=0,  # 客户创建时金额为0，后续填写
                sign_date=timezone.now().date(),  # 使用当前日期作为签约日期
                end_date=None,  # 待签约，未设置结束日期
                status='pending',  # 待签约状态
                create_user_id=user.id,
                auto_generated=True  # 标记为自动生成
            )
        except Exception as e:
            logger.error(f"创建CustomerContract记录失败: {e}")
        
        # 2. 生成Contract记录（销售合同）
        try:
            from apps.contract.models import Contract
            # 创建Contract记录
            contract = Contract.objects.create(
                customer_id=customer.id,
                customer=customer.name,
                code=f"CONTRACT-{customer.id}-{int(time.time())}",
                name=f"{customer.name}销售合同",
                cate_id=1,  # 默认分类ID
                types=1,  # 普通合同
                admin_id=user.id,
                prepared_uid=user.id,
                cost=0.00,  # 客户创建时金额为0，后续填写
                check_status=0,  # 待审核状态
                delete_time=0,  # 未删除
                auto_generated=True  # 标记为自动生成
            )
        except Exception as e:
            logger.error(f"创建Contract记录失败: {e}")
        
        # 2. 生成待确认订单记录
        try:
            from apps.customer.models import CustomerOrder
            # 创建客户订单记录（待确认）
            order_record = CustomerOrder.objects.create(
                customer_id=customer.id,
                order_number=f"ORD-CUST-{customer.id}-{int(time.time())}",
                product_name=f"{customer.name}相关产品",
                amount=0,  # 客户创建时金额为0，后续填写
                order_date=timezone.now().date(),
                status='pending',  # 待处理状态
                description=f"客户{customer.name}相关订单",
                create_user_id=user.id,
                auto_generated=True  # 标记为自动生成
            )
        except Exception as e:
            logger.error(f"创建客户订单记录失败: {e}")
        
        # 3. 生成项目记录（因为发票记录需要项目ID）
        project_record = None
        try:
            from apps.project.models import Project
            # 创建项目记录
            project_record = Project.objects.create(
                customer_id=customer.id,
                name=f"{customer.name}项目",
                code=f"PROJ-CUST-{customer.id}-{int(time.time())}",
                budget=0,  # 客户创建时预算为0，后续填写
                start_date=None,  # 待开始，未设置开始日期
                end_date=None,  # 待结束，未设置结束日期
                status=1,  # 未开始状态
                creator=user,
                auto_generated=True  # 标记为自动生成
            )
        except Exception as e:
            logger.error(f"创建项目记录失败: {e}")
        
        # 4. 生成待收款记录（需要项目ID）
        try:
            from apps.finance.models import Invoice
            # 创建发票记录（待收款）- 确保提供项目ID
            invoice_data = {
                'code': f"INV-CUST-{customer.id}-{int(time.time())}",
                'customer_id': customer.id,
                'amount': 0,  # 客户创建时金额为0，后续填写
                'admin_id': user.id,
                'did': 1,  # 默认部门ID
                'open_status': 0,  # 未开票状态
                'enter_status': 0,  # 未回款状态
                'invoice_title': "",  # 留空，后续填写
                'invoice_tax': "",  # 留空，后续填写
                'create_time': int(time.time())
            }
            
            # 如果项目记录创建成功，添加项目ID
            if project_record and hasattr(project_record, 'id'):
                invoice_data['project_id'] = project_record.id
            
            invoice_record = Invoice.objects.create(**invoice_data)
        except Exception as e:
            logger.error(f"创建发票记录失败: {e}")
    
    def form_invalid(self, form):
        messages.error(self.request, '表单验证失败，请检查输入信息')
        return super().form_invalid(form)
    
    def save_custom_fields(self):
        """保存自定义字段值"""
        custom_fields = CustomerField.objects.filter(status=True, delete_time=0)
        
        for field in custom_fields:
            field_key = f'custom_field_{field.id}'
            field_value = self.request.POST.get(field_key, '')
            
            if field.field_type == 'checkbox':
                field_value = '1' if field_value else '0'
            
            # 创建或更新字段值
            CustomerCustomFieldValue.objects.update_or_create(
                customer=self.object,
                field=field,
                defaults={'value': field_value}
            )


class CustomerUpdateView(LoginRequiredMixin, UpdateView):
    form_class = CustomerForm
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = Customer
    template_name = 'customer/customer_form.html'
    success_url = reverse_lazy('customer:customer_list')
    
    def get_queryset(self):
        queryset = Customer.objects.filter(delete_time=0)
        
        # 添加数据权限过滤
        user = self.request.user
        
        # 超级管理员可以查看所有客户
        if hasattr(user, 'is_superuser') and user.is_superuser:
            return queryset
        
        # 数据权限过滤：只能查看自己的客户及共享给自己的客户
        queryset = queryset.filter(
            models.Q(belong_uid=user.id) | 
            models.Q(share_ids__contains=str(user.id))
        )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # 获取启用的自定义字段
        custom_fields = CustomerField.objects.filter(status=True, delete_time=0).order_by('sort')
        
        # 获取当前客户的自定义字段值
        current_values = {}
        for cfv in CustomerCustomFieldValue.objects.filter(customer=self.object):
            current_values[cfv.field_id] = cfv.value
        
        # 为每个字段添加选项列表和当前值
        for field in custom_fields:
            if field.options:
                field.options_list = [opt.strip() for opt in field.options.split('\n') if opt.strip()]
            else:
                field.options_list = []
            field.current_value = current_values.get(field.id, '')
        
        context['custom_fields'] = custom_fields
        
        # 联系人表单集
        if self.request.POST:
            context['contact_formset'] = ContactFormSet(self.request.POST, instance=self.object)
        else:
            context['contact_formset'] = ContactFormSet(instance=self.object)
        
        return context
    
    def form_valid(self, form):
        context = self.get_context_data()
        contact_formset = context['contact_formset']
        
        with transaction.atomic():
            self.object = form.save()
            
            # 处理主要联系人的单选逻辑
            primary_contact_value = self.request.POST.get('primary_contact')
            
            # 保存联系人信息
            if contact_formset.is_valid():
                contact_formset.instance = self.object
                contacts = contact_formset.save(commit=False)
                
                # 处理删除标记的联系人
                for contact in contact_formset.deleted_objects:
                    contact.delete()
                
                # 保存新建和修改的联系人
                for contact in contacts:
                    contact.save()
                
                # 处理主要联系人设置
                if primary_contact_value:
                    # 先将所有联系人设为非主要
                    Contact.objects.filter(customer=self.object).update(is_primary=False)
                    
                    # 根据primary_contact的值设置主要联系人
                    try:
                        contact_id = int(primary_contact_value)
                        Contact.objects.filter(id=contact_id, customer=self.object).update(is_primary=True)
                    except (ValueError, TypeError):
                        # 如果转换失败，可能是新添加的联系人索引
                        if primary_contact_value.isdigit():
                            contact_index = int(primary_contact_value)
                            if contact_index < len(contacts):
                                contacts[contact_index].is_primary = True
                                contacts[contact_index].save()
            else:
                return self.form_invalid(form)
            
            # 处理自定义字段
            self.save_custom_fields()
            
            # 添加操作日志
            SystemLog.objects.create(
                user=self.request.user,
                log_type='update',
                module='客户管理',
                action='更新客户',
                content=f'成功更新客户: {self.object.name}',
                ip_address=self.request.META.get('REMOTE_ADDR'),
                user_agent=self.request.META.get('HTTP_USER_AGENT')
            )
        
        messages.success(self.request, '客户信息更新成功！')
        # 直接使用reverse生成URL进行重定向，确保重定向正确执行
        from django.urls import reverse
        return HttpResponseRedirect(reverse('customer:customer_list'))
    
    def form_invalid(self, form):
        messages.error(self.request, '表单验证失败，请检查输入信息')
        return super().form_invalid(form)
    
    def save_custom_fields(self):
        """保存自定义字段值"""
        custom_fields = CustomerField.objects.filter(status=True, delete_time=0)
        
        for field in custom_fields:
            field_key = f'custom_field_{field.id}'
            field_value = self.request.POST.get(field_key, '')
            
            if field.field_type == 'checkbox':
                field_value = '1' if field_value else '0'
            
            # 创建或更新字段值
            CustomerCustomFieldValue.objects.update_or_create(
                customer=self.object,
                field=field,
                defaults={'value': field_value}
            )


class CustomerDetailView(LoginRequiredMixin, DetailView):
    model = Customer
    template_name = 'customer/customer_detail.html'
    context_object_name = 'customer'
    login_url = '/user/login/'
    redirect_field_name = 'next'
    
    def get_queryset(self):
        # 基础过滤：只返回未删除的客户
        queryset = Customer.objects.filter(delete_time=0)
        
        # 添加数据权限过滤
        user = self.request.user
        
        # 超级管理员可以查看所有客户
        if hasattr(user, 'is_superuser') and user.is_superuser:
            return queryset
        
        # 数据权限过滤：只能查看自己的客户及共享给自己的客户
        queryset = queryset.filter(
            models.Q(belong_uid=user.id) | 
            models.Q(share_ids__contains=str(user.id))
        )
        
        return queryset
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 获取客户的自定义字段值
        custom_field_values = {}
        for cfv in CustomerCustomFieldValue.objects.filter(customer=self.object).select_related('field'):
            custom_field_values[cfv.field.name] = cfv.value
        
        # 获取基础数据映射
        try:
            customer_grade = CustomerGrade.objects.filter(id=self.object.grade_id, delete_time=0).first()
            customer_intent = CustomerIntent.objects.filter(id=self.object.services_id, delete_time=0).first()
        except:
            customer_grade = None
            customer_intent = None
        
        context['custom_field_values'] = custom_field_values
        context['customer_grade'] = customer_grade
        context['customer_intent'] = customer_intent
        context['contacts'] = self.object.contacts.all()
        context['follow_records'] = self.object.follow_records.all()[:10]  # 最近10条跟进记录
        context['orders'] = self.object.orders.filter(delete_time=0)[:5]  # 最近5个订单
        
        # 获取项目标准合同模块的合同数据，而不是独立的CustomerContract数据
        try:
            from apps.contract.models import Contract
            context['contracts'] = Contract.objects.filter(customer_id=self.object.id, delete_time=0)[:5]  # 最近5个合同
        except Exception as e:
            logger.error(f"获取标准合同数据失败: {str(e)}")
            context['contracts'] = []
        
        # 获取客户发票记录 - 整合客户模块和财务模块的发票记录
        try:
            from apps.finance_new.models import Invoice
            # 获取客户模块的发票记录（使用SoftDeleteModel的软删除机制）
            customer_invoices = Invoice.objects.filter(customer_id=self.object.id, is_deleted=False)[:5]
            
            # 获取财务模块的开票记录（关联到该客户的发票）
            finance_invoices = Invoice.objects.filter(customer_id=self.object.id, is_deleted=False).order_by('-id')[:10]
            
            # 合并发票记录，按开票时间排序
            all_invoices = list(customer_invoices) + list(finance_invoices)
            # 去重并排序
            seen_ids = set()
            unique_invoices = []
            for invoice in all_invoices:
                if invoice.id not in seen_ids:
                    seen_ids.add(invoice.id)
                    unique_invoices.append(invoice)
            
            # 按开票时间排序
            unique_invoices.sort(key=lambda x: x.issued_at if hasattr(x, 'issued_at') and x.issued_at else x.created_at, reverse=True)
            
            context['invoices'] = unique_invoices[:10]  # 显示最多10张发票
        except Exception as e:
            logger.error(f"获取发票记录失败: {str(e)}")
            context['invoices'] = []
        
        # 获取客户财务往来记录 - 直接列表显示（按日期排序）
        try:
            # 获取所有发票记录 - 使用models_new中的Invoice模型和正确的软删除字段
            from apps.finance_new.models import Invoice, Income, Payment
            all_invoices = Invoice.objects.filter(customer=self.object.id, is_deleted=False).order_by('-id')[:20]
            
            # 获取所有收款记录（Income模型）- 通过invoice关联到customer
            all_incomes = Income.objects.filter(invoice__customer=self.object.id).order_by('-id')[:20]
            
            # 获取所有付款记录（Payment模型）- 暂时不显示付款记录，因为Payment模型与客户没有直接关联
            all_payments = Payment.objects.none()  # 返回空查询集
            
            # 构建完整的财务记录列表
            financial_records = []
            
            # 处理发票记录
            for invoice in all_invoices:
                invoice_date = None
                try:
                    if hasattr(invoice, 'open_time') and invoice.open_time > 0:
                        invoice_date = datetime.fromtimestamp(invoice.open_time)
                    elif hasattr(invoice, 'create_time'):
                        invoice_date = datetime.fromtimestamp(invoice.create_time) if isinstance(invoice.create_time, (int, float)) else invoice.create_time
                except:
                    pass
                
                if invoice_date:
                    financial_records.append({
                        'type': 'invoice',
                        'record': invoice,
                        'amount': invoice.amount,
                        'date': invoice_date,
                        'status': '已开票' if invoice.open_status == 1 else '未开票'
                    })
            
            # 处理收款记录
            for income in all_incomes:
                income_date = None
                try:
                    if hasattr(income, 'income_date'):
                        income_date = income.income_date
                    elif hasattr(income, 'create_time'):
                        income_date = datetime.fromtimestamp(income.create_time) if isinstance(income.create_time, (int, float)) else income.create_time
                except:
                    pass
                
                if income_date:
                    financial_records.append({
                        'type': 'income',
                        'record': income,
                        'amount': income.amount,
                        'date': income_date,
                        'status': '已收款'
                    })
            
            # 处理付款记录
            for payment in all_payments:
                payment_date = None
                try:
                    if hasattr(payment, 'payment_date'):
                        payment_date = payment.payment_date
                    elif hasattr(payment, 'create_time'):
                        payment_date = datetime.fromtimestamp(payment.create_time) if isinstance(payment.create_time, (int, float)) else payment.create_time
                except:
                    pass
                
                if payment_date:
                    financial_records.append({
                        'type': 'payment',
                        'record': payment,
                        'amount': payment.amount,
                        'date': payment_date,
                        'status': '已付款'
                    })
            
            # 按日期排序
            financial_records.sort(key=lambda x: x['date'], reverse=True)
            
            # 直接提供按日期排序的财务记录列表
            context['financial_records'] = financial_records[:20]  # 按日期排序的财务记录列表
            
        except Exception as e:
            logger.error(f"获取财务记录失败: {str(e)}")
            context['financial_records'] = []
        
        # 获取客户项目记录
        try:
            from apps.project.models import Project
            context['projects'] = Project.objects.filter(customer_id=self.object.id, delete_time__isnull=True)[:5]  # 最近5个项目
        except:
            context['projects'] = []
        
        return context


class CustomerDetailApiView(LoginRequiredMixin, View):
    login_url = '/user/login/'
    redirect_field_name = 'next'
    
    def get(self, request, pk):
        try:
            customer = get_object_or_404(Customer, id=pk, delete_time=0)
            
            # 获取主要联系人信息
            primary_contact = customer.contacts.filter(is_primary=True).first()
            contact_person = primary_contact.contact_person if primary_contact else ''
            contact_phone = primary_contact.phone if primary_contact else ''
            
            data = {
                'id': customer.id,
                'name': customer.name,
                'contact_person': contact_person,
                'phone': contact_phone or customer.phone or '',
                'address': customer.address or '',
            }
            
            return JsonResponse(data, json_dumps_params={'ensure_ascii': False})
        except Exception as e:
            logger.error(f'获取客户详情API失败: {str(e)}', exc_info=True)
            return JsonResponse({'error': str(e)}, status=404, json_dumps_params={'ensure_ascii': False})


class CustomerDeleteView(LoginRequiredMixin, View):
    login_url = '/user/login/'
    redirect_field_name = 'next'
    
    def post(self, request, pk):
        try:
            customer = Customer.objects.get(id=pk, delete_time=0)
            
            # 检查当前用户是否有删除权限
            # 1. 超级管理员可以删除任何客户
            # 2. 客户归属人可以删除自己的客户
            # 3. 共享用户可以删除共享给自己的客户
            if not (hasattr(request.user, 'is_superuser') and request.user.is_superuser):
                # 检查是否是归属人
                is_owner = customer.belong_uid == request.user.id
                
                # 检查是否是共享用户
                is_shared = False
                if customer.share_ids:
                    shared_ids = customer.share_ids.split(',')
                    is_shared = str(request.user.id) in shared_ids
                
                if not (is_owner or is_shared):
                    return JsonResponse({'status': 'error', 'msg': '您没有权限删除此客户'}, status=403)
            
            # 保存客户名称用于日志记录
            customer_name = customer.name
            
            # 客户从个人列表删除后自动流转至公海：
            # 1. 清除归属人ID
            # 2. 清除部门ID
            # 3. 清除共享ID列表，确保不会再显示在原归属人的客户列表中
            # 4. 保留其他信息
            customer.belong_uid = 0
            customer.belong_did = 0
            customer.share_ids = ''
            customer.save()
            
            # 添加操作日志
            SystemLog.objects.create(
                user=request.user,
                log_type='update',
                module='客户管理',
                action='客户移入公海',
                content=f'将客户 "{customer_name}" 移入公海',
                ip_address=request.META.get('REMOTE_ADDR'),
                user_agent=request.META.get('HTTP_USER_AGENT')
            )
            
            return JsonResponse({'status': 'success', 'msg': '客户已成功移入公海'})
        except Customer.DoesNotExist:
            return JsonResponse({'status': 'error', 'msg': '客户不存在'}, status=404)
        except Exception as e:
            logger.error(f"删除客户失败: {str(e)}")
            return JsonResponse({'status': 'error', 'msg': f'操作失败: {str(e)}'}, status=500)


class CustomerBatchDeleteView(LoginRequiredMixin, View):
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def post(self, request):
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])
            if not ids:
                return JsonResponse({'status': 'error', 'message': '未提供客户ID列表'}, status=400)

            # 获取当前用户可操作的客户
            if hasattr(request.user, 'is_superuser') and request.user.is_superuser:
                # 超级管理员可以操作所有客户
                customers = Customer.objects.filter(id__in=ids, delete_time=0)
            else:
                # 普通用户只能操作自己的客户或共享给自己的客户
                customers = Customer.objects.filter(
                    Q(id__in=ids) &
                    Q(delete_time=0) &
                    (Q(belong_uid=request.user.id) | 
                     Q(share_ids__contains=str(request.user.id)))
                )

            # 批量将客户移入公海：
            # 1. 清除归属人ID
            # 2. 清除部门ID
            # 3. 清除共享ID列表
            updated_count = customers.update(
                belong_uid=0,
                belong_did=0,
                share_ids=''
            )

            # 添加操作日志
            if updated_count > 0:
                customer_names = ', '.join(customers.values_list('name', flat=True))
                SystemLog.objects.create(
                    user=request.user,
                    log_type='delete',
                    module='customer',
                    action='批量删除客户',
                    content=f'成功将{updated_count}个客户移入公海: {customer_names}',
                    ip_address=request.META.get('REMOTE_ADDR', '0.0.0.0'),
                    user_agent=request.META.get('HTTP_USER_AGENT', '')
                )

            return JsonResponse({'status': 'success', 'message': f'成功将{updated_count}个客户移入公海'})
        except Exception as e:
            logger.error(f'批量处理客户失败: {str(e)}')
            return JsonResponse({'status': 'error', 'message': f'操作失败: {str(e)}'}, status=500)


class CustomerBatchImportView(LoginRequiredMixin, TemplateView):
    template_name = 'customer/customer_import.html'
    login_url = '/user/login/'
    redirect_field_name = 'next'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        
        # 获取基础数据选项，用于导入时的数据映射
        context['customer_sources'] = CustomerSource.objects.filter(status=1, delete_time=0).order_by('sort', 'id')
        context['customer_grades'] = CustomerGrade.objects.filter(status=1, delete_time=0).order_by('sort', 'id')
        context['customer_intents'] = CustomerIntent.objects.filter(status=1, delete_time=0).order_by('sort', 'id')
        context['custom_fields'] = CustomerField.objects.filter(status=True, delete_time=0).order_by('sort', 'id')
        
        return context


# 添加缺失的视图函数
@login_required
# 客户字段管理相关视图将在文件下方定义



@login_required
def customer_field_create(request):
    """创建客户字段"""
    if request.method == 'POST':
        try:
            name = request.POST.get('name', '').strip()
            field_type = request.POST.get('field_type', '')
            options = request.POST.get('options', '').strip()
            is_required = request.POST.get('is_required') == 'on'
            is_unique = request.POST.get('is_unique') == 'on'
            is_list_display = request.POST.get('is_list_display') == 'on'
            sort = int(request.POST.get('sort', 0))
            
            if not name or not field_type:
                return JsonResponse({'code': 1, 'msg': '字段名称和类型不能为空'})
            
            # 生成字段标识
            field_name = generate_field_name(name)
            
            # 检查字段标识是否重复
            if CustomerField.objects.filter(field_name=field_name, delete_time=0).exists():
                field_name = f"{field_name}_{int(timezone.now().timestamp())}"
            
            # 创建字段
            CustomerField.objects.create(
                name=name,
                field_name=field_name,
                field_type=field_type,
                options=options,
                is_required=is_required,
                is_unique=is_unique,
                is_list_display=is_list_display,
                sort=sort,
                status=True,
                delete_time=0
            )
            
            return JsonResponse({'code': 0, 'msg': '创建成功'}, json_dumps_params={'ensure_ascii': False})
            
        except Exception as e:
            logger.error(f"创建客户字段失败: {str(e)}")
            return JsonResponse({'code': 1, 'msg': f'创建失败: {str(e)}'}, json_dumps_params={'ensure_ascii': False})
    
    return render(request, 'customer/customer_field_form.html')


@login_required
def customer_field_edit(request, pk):
    """编辑客户字段"""
    try:
        field = get_object_or_404(CustomerField, pk=pk, delete_time=0)
        
        if request.method == 'POST':
            field.name = request.POST.get('name', '').strip()
            field.field_type = request.POST.get('field_type', '')
            field.options = request.POST.get('options', '').strip()
            field.is_required = request.POST.get('is_required') == 'on'
            field.is_unique = request.POST.get('is_unique') == 'on'
            field.is_list_display = request.POST.get('is_list_display') == 'on'
            field.sort = int(request.POST.get('sort', 0))
            
            if not field.name or not field.field_type:
                return JsonResponse({'code': 1, 'msg': '字段名称和类型不能为空'})
            
            field.save()
            return JsonResponse({'code': 0, 'msg': '更新成功'})
        
        # 创建表单对象
        form = CustomerFieldForm(instance=field)
        return render(request, 'basedata/customer/customer_field_form.html', {'object': field, 'form': form})
        
    except Exception as e:
        logger.error(f"编辑客户字段失败: {str(e)}")
        return JsonResponse({'code': 1, 'msg': f'编辑失败: {str(e)}'})


@login_required
def customer_field_delete(request, pk):
    """删除客户字段"""
    try:
        field = get_object_or_404(CustomerField, pk=pk, delete_time=0)
        field.delete_time = int(timezone.now().timestamp())
        field.save()
        
        return JsonResponse({'code': 0, 'msg': '删除成功'})
        
    except Exception as e:
        logger.error(f"删除客户字段失败: {str(e)}")
        return JsonResponse({'code': 1, 'msg': f'删除失败: {str(e)}'})


@login_required
def customer_field_toggle(request, pk):
    """切换客户字段状态"""
    try:
        field = get_object_or_404(CustomerField, pk=pk, delete_time=0)
        field.status = not field.status
        field.save()
        
        status_text = '启用' if field.status else '禁用'
        return JsonResponse({'code': 0, 'msg': f'{status_text}成功'})
        
    except Exception as e:
        logger.error(f"切换客户字段状态失败: {str(e)}")
        return JsonResponse({'code': 1, 'msg': f'操作失败: {str(e)}'})


def generate_field_name(name):
    """生成字段标识"""
    import re
    
    # 中文转拼音映射表（简化版）
    pinyin_map = {
        '客': 'ke', '户': 'hu', '账': 'zhang', '号': 'hao',
        '姓': 'xing', '名': 'ming', '电': 'dian', '话': 'hua',
        '手': 'shou', '机': 'ji', '邮': 'you', '箱': 'xiang',
        '地': 'di', '址': 'zhi', '公': 'gong', '司': 'si',
        '联': 'lian', '系': 'xi', '人': 'ren', '备': 'bei',
        '注': 'zhu', '描': 'miao', '述': 'shu', '类': 'lei',
        '型': 'xing', '状': 'zhuang', '态': 'tai'
    }
    
    result = ''
    for char in name:
        if char in pinyin_map:
            result += pinyin_map[char]
        elif char.isalnum():
            result += char.lower()
    
    # 清理结果
    result = re.sub(r'[^a-zA-Z0-9]', '', result)
    return result or 'custom_field'
# 公海客户管理视图
class PublicCustomerListView(LoginRequiredMixin, ListView):
    """公海客户列表视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = Customer
    template_name = 'customer/public_customer_list.html'
    context_object_name = 'customers'

    def get_queryset(self):
        # 公海客户：没有归属人且未废弃的客户
        return Customer.objects.filter(delete_time=0, belong_uid=0, discard_time=0)

class PublicCustomerListDataView(LoginRequiredMixin, View):
    """公海客户列表数据API"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def get(self, request):
        try:
            # 公海客户：没有归属人且未废弃的客户
            queryset = Customer.objects.filter(delete_time=0, belong_uid=0, discard_time=0)
            
            # 处理搜索
            search = request.GET.get('search', '')
            if search:
                queryset = queryset.filter(name__icontains=search)
            
            # 分页
            page = int(request.GET.get('page', 1))
            limit = int(request.GET.get('limit', 20))
            start = (page - 1) * limit
            end = start + limit
            
            total_count = queryset.count()
            paginated_queryset = queryset[start:end]
            
            data = []
            for customer in paginated_queryset:
                # 获取客户等级名称
                grade_name = ''
                if customer.grade_id > 0:
                    try:
                        from apps.customer.models import CustomerGrade
                        grade = CustomerGrade.objects.filter(id=customer.grade_id).first()
                        grade_name = grade.title if grade else ''
                    except:
                        pass
                
                # 获取客户来源名称
                source_name = customer.customer_source.title if customer.customer_source else ''
                
                # 获取主要联系人
                from apps.customer.models import Contact
                primary_contact = Contact.objects.filter(customer=customer, is_primary=True).first()
                contact_name = primary_contact.contact_person if primary_contact else ''
                contact_phone = primary_contact.phone if primary_contact else ''
                
                data.append({
                    'id': customer.id,
                    'name': customer.name,
                    'province': customer.province,
                    'city': customer.city,
                    'district': customer.district,
                    'address': customer.address,
                    'grade': grade_name,
                    'grade_id': customer.grade_id,
                    'source': source_name,
                    'content': customer.content[:50] + '...' if customer.content and len(customer.content) > 50 else customer.content,
                    'contact_name': contact_name,
                    'contact_phone': contact_phone,
                    'market': customer.market[:50] + '...' if customer.market and len(customer.market) > 50 else customer.market,
                    'remark': customer.remark[:50] + '...' if customer.remark and len(customer.remark) > 50 else customer.remark,
                    'create_time': customer.create_time.strftime('%Y-%m-%d %H:%M:%S') if customer.create_time else '',
                    'tax_num': customer.tax_num,
                })
            
            return JsonResponse({
                'code': 0,
                'msg': '',
                'count': total_count,
                'data': data
            })
        except Exception as e:
            return JsonResponse({'code': 1, 'msg': f'获取失败: {str(e)}'})

# 爬虫任务管理视图
class SpiderTaskListView(LoginRequiredMixin, ListView):
    """爬虫任务列表视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = SpiderTask
    template_name = 'customer/spider_task_list.html'
    context_object_name = 'tasks'

class SpiderTaskListDataView(LoginRequiredMixin, View):
    """爬虫任务列表数据API"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def get(self, request):
        try:
            tasks = SpiderTask.objects.filter(delete_time=0).order_by('-create_time')
            
            # 处理搜索
            search = request.GET.get('search', '')
            if search:
                tasks = tasks.filter(task_name__icontains=search)
            
            # 分页
            page = int(request.GET.get('page', 1))
            limit = int(request.GET.get('limit', 20))
            start = (page - 1) * limit
            end = start + limit
            
            total_count = tasks.count()
            paginated_tasks = tasks[start:end]
            
            data = []
            for task in paginated_tasks:
                data.append({
                    'id': task.id,
                    'task_name': task.task_name,
                    'spider_keywords': task.spider_keywords,
                    'status': task.status,
                    'status_display': task.get_status_display(),
                    'create_time': task.create_time.strftime('%Y-%m-%d %H:%M:%S'),
                    'create_user': task.create_user.username if task.create_user else ''
                })
            
            return JsonResponse({
                'code': 0,
                'msg': '',
                'count': total_count,
                'data': data
            })
        except Exception as e:
            return JsonResponse({'code': 1, 'msg': f'获取失败: {str(e)}'})

class SpiderTaskCreateView(LoginRequiredMixin, CreateView):
    """创建爬虫任务视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = SpiderTask
    template_name = 'customer/spider_task_form.html'
    fields = ['task_name', 'spider_keywords', 'data_region', 'industry_limit', 'province', 'insured_count', 'contact_phone']
    success_url = reverse_lazy('customer:spider_task_list')

    def form_valid(self, form):
        form.instance.create_user = self.request.user
        return super().form_valid(form)

class SpiderTaskUpdateView(LoginRequiredMixin, UpdateView):
    """编辑爬虫任务视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = SpiderTask
    template_name = 'customer/spider_task_form.html'
    fields = ['task_name', 'spider_keywords', 'data_region', 'industry_limit', 'province', 'insured_count', 'contact_phone']
    success_url = reverse_lazy('customer:spider_task_list')

class SpiderTaskDeleteView(LoginRequiredMixin, DeleteView):
    """删除爬虫任务视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = SpiderTask
    success_url = reverse_lazy('customer:spider_task_list')

class SpiderTaskActionView(LoginRequiredMixin, View):
    """爬虫任务操作视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def post(self, request, pk):
        try:
            task = SpiderTask.objects.get(id=pk, delete_time=0)
            action = request.POST.get('action')
            
            if action == 'start':
                task.status = 1  # 运行中
                task.save()
                return JsonResponse({'code': 0, 'msg': '任务已启动'})
            elif action == 'stop':
                task.status = 2  # 已停止
                task.save()
                return JsonResponse({'code': 0, 'msg': '任务已停止'})
            else:
                return JsonResponse({'code': 1, 'msg': '无效的操作'})
                
        except SpiderTask.DoesNotExist:
            return JsonResponse({'code': 1, 'msg': '任务不存在'})
        except Exception as e:
            return JsonResponse({'code': 1, 'msg': f'操作失败: {str(e)}'})

# 获取员工列表视图
@login_required
def get_employee_list(request):
    """
    获取员工列表，用于客户共享选择
    """
    try:
        # 获取所有员工
        from apps.user.models import Admin
        employees = Admin.objects.filter(status=1).values('id', 'username', 'name')
        return JsonResponse(list(employees), safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# 获取已共享员工视图
@login_required
def get_shared_employees(request, customer_id):
    """
    获取指定客户已共享的员工ID列表
    """
    try:
        customer = Customer.objects.get(id=customer_id, delete_time=0)
        # 解析share_ids字段，获取已共享的员工ID列表
        shared_ids = customer.share_ids.split(',') if customer.share_ids else []
        # 过滤空字符串并转换为整数
        shared_ids = [int(id) for id in shared_ids if id]
        return JsonResponse(shared_ids, safe=False)
    except Customer.DoesNotExist:
        return JsonResponse([], safe=False)
    except Exception as e:
        return JsonResponse({'error': str(e)}, status=500)

# 设置客户共享视图
@login_required
def set_customer_share(request, customer_id):
    """
    设置客户共享给指定员工
    """
    try:
        # 获取当前用户
        user = request.user
        # 获取客户对象
        customer = get_object_or_404(Customer, id=customer_id, delete_time=0)
        
        # 检查权限：只有客户归属人可以设置共享
        if customer.belong_uid != user.id and not (hasattr(user, 'is_superuser') and user.is_superuser):
            return JsonResponse({'status': 'error', 'msg': '您没有权限设置此客户的共享'}, status=403)
        
        # 获取要共享的员工ID列表
        shared_user_ids = request.POST.getlist('user_ids[]', [])
        # 过滤并转换为整数
        shared_user_ids = [int(id) for id in shared_user_ids if id.isdigit()]
        
        # 转换为逗号分隔的字符串
        share_ids_str = ','.join(map(str, shared_user_ids))
        
        # 更新客户的共享字段
        customer.share_ids = share_ids_str
        customer.save()
        
        # 添加操作日志
        SystemLog.objects.create(
            user=request.user,
            log_type='update',
            module='客户管理',
            action='设置客户共享',
            content=f'将客户 "{customer.name}" 共享给员工ID: {share_ids_str}',
            ip_address=request.META.get('REMOTE_ADDR'),
            user_agent=request.META.get('HTTP_USER_AGENT')
        )
        
        return JsonResponse({'status': 'success', 'msg': '客户共享设置成功'})
    except Exception as e:
        logger.error(f"设置客户共享失败: {str(e)}")
        return JsonResponse({'status': 'error', 'msg': f'操作失败: {str(e)}'}, status=500)

# 公海客户认领视图
@login_required
def claim_public_customer(request, customer_id):
    """
    认领公海客户
    """
    try:
        customer = Customer.objects.get(id=customer_id, delete_time=0)
        
        # 检查客户是否在公海（没有归属人）
        if customer.belong_uid != 0:
            return JsonResponse({'status': 'error', 'msg': '该客户已被认领'})
        
        # 设置客户归属人为当前用户
        customer.belong_uid = request.user.id
        customer.belong_did = request.user.main_department_id if hasattr(request.user, 'main_department_id') else 0
        customer.belong_time = int(timezone.now().timestamp())
        customer.save()
        
        return JsonResponse({'status': 'success', 'msg': '客户认领成功'})
    except Customer.DoesNotExist:
        return JsonResponse({'status': 'error', 'msg': '客户不存在'})
    except Exception as e:
        logger.error(f'认领客户失败: {str(e)}')
        return JsonResponse({'status': 'error', 'msg': f'认领失败: {str(e)}'})

# 客户移入废弃列表视图
@login_required
def discard_customer(request, customer_id):
    """
    将客户移入废弃列表
    """
    try:
        customer = Customer.objects.get(id=customer_id, delete_time=0)
        
        # 将客户移入废弃列表：
        # 1. 设置discard_time为当前时间戳
        # 2. 保留其他信息
        customer.discard_time = int(timezone.now().timestamp())
        customer.save()
        
        return JsonResponse({'status': 'success', 'msg': '客户已成功移入废弃列表'})
    except Customer.DoesNotExist:
        return JsonResponse({'status': 'error', 'msg': '客户不存在'})
    except Exception as e:
        logger.error(f'移入废弃列表失败: {str(e)}')
        return JsonResponse({'status': 'error', 'msg': f'操作失败: {str(e)}'})

# 废弃客户恢复视图
@login_required
def restore_customer(request, customer_id):
    """
    将废弃客户恢复至公海
    """
    try:
        customer = Customer.objects.get(id=customer_id, delete_time=0)
        
        # 将废弃客户恢复至公海：
        # 1. 清除discard_time
        # 2. 清除归属人ID和部门ID（恢复到公海）
        customer.discard_time = 0
        customer.belong_uid = 0
        customer.belong_did = 0
        customer.save()
        
        return JsonResponse({'status': 'success', 'msg': '客户已成功恢复至公海'})
    except Customer.DoesNotExist:
        return JsonResponse({'status': 'error', 'msg': '客户不存在'})
    except Exception as e:
        logger.error(f'恢复客户失败: {str(e)}')
        return JsonResponse({'status': 'error', 'msg': f'操作失败: {str(e)}'})

# 废弃客户清理视图
@login_required
def clean_abandoned_customers(request):
    """
    清理过期废弃客户
    """
    try:
        # 获取过期时间（默认90天）
        days = int(request.GET.get('days', 90))
        cutoff_time = int((timezone.now() - timedelta(days=days)).timestamp())
        
        # 清理过期废弃客户：
        # 1. 只清理discard_time大于0且超过过期时间的客户
        # 2. 使用软删除，设置delete_time为当前时间戳
        deleted_count = Customer.objects.filter(
            delete_time=0,
            discard_time__gt=0,
            discard_time__lt=cutoff_time
        ).update(
            delete_time=int(timezone.now().timestamp())
        )
        
        return JsonResponse({'status': 'success', 'msg': f'成功清理{deleted_count}个过期废弃客户'})
    except Exception as e:
        logger.error(f'清理废弃客户失败: {str(e)}')
        return JsonResponse({'status': 'error', 'msg': f'操作失败: {str(e)}'})

# 自动将长期未跟进的客户移入公海
@login_required
def auto_move_to_public_pool(request):
    """
    自动将长期未跟进的客户移入公海
    """
    try:
        # 获取未跟进天数（默认30天）
        days = int(request.GET.get('days', 30))
        cutoff_time = int((timezone.now() - timedelta(days=days)).timestamp())
        
        # 查找符合条件的客户：
        # 1. 非删除状态
        # 2. 有归属人（belong_uid > 0）
        # 3. 最近跟进时间follow_time < cutoff_time
        # 4. 未废弃（discard_time = 0）
        customers_to_move = Customer.objects.filter(
            delete_time=0,
            belong_uid__gt=0,
            follow_time__lt=cutoff_time,
            discard_time=0
        )
        
        # 记录要移动的客户数量
        move_count = customers_to_move.count()
        
        if move_count > 0:
            # 开始事务
            with transaction.atomic():
                # 批量更新客户信息
                customers_to_move.update(
                    belong_uid=0,  # 移除归属人
                    belong_did=0,  # 移除归属部门
                    share_ids='',  # 清空共享
                    distribute_time=int(timezone.now().timestamp())  # 更新分配时间
                )
                
                # 记录操作日志
                for customer in customers_to_move:
                    SystemLog.objects.create(
                        user=request.user,
                        log_type='update',
                        module='客户管理',
                        action='自动移入公海',
                        content=f'客户 "{customer.name}" 因超过{days}天未跟进，自动移入公海',
                        ip_address=request.META.get('REMOTE_ADDR'),
                        user_agent=request.META.get('HTTP_USER_AGENT')
                    )
        
        return JsonResponse({'status': 'success', 'msg': f'成功将{move_count}个长期未跟进客户移入公海'})
    except Exception as e:
        logger.error(f'自动移入公海失败: {str(e)}')
        return JsonResponse({'status': 'error', 'msg': f'操作失败: {str(e)}'})

# 废弃客户管理视图
class AbandonedCustomerListView(LoginRequiredMixin, ListView):
    """废弃客户列表视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = Customer
    template_name = 'customer/abandoned_customer_list.html'
    context_object_name = 'customers'

    def get_queryset(self):
        # 废弃客户：discard_time > 0的客户
        return Customer.objects.filter(delete_time=0).exclude(discard_time=0)

class AbandonedCustomerListDataView(LoginRequiredMixin, View):
    """废弃客户列表数据API"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def get(self, request):
        try:
            queryset = Customer.objects.filter(delete_time=0).exclude(discard_time=0)
            
            # 处理搜索
            search = request.GET.get('search', '')
            if search:
                queryset = queryset.filter(name__icontains=search)
            
            # 分页
            page = int(request.GET.get('page', 1))
            limit = int(request.GET.get('limit', 20))
            start = (page - 1) * limit
            end = start + limit
            
            total_count = queryset.count()
            paginated_queryset = queryset[start:end]
            
            data = []
            for customer in paginated_queryset:
                data.append({
                    'id': customer.id,
                    'name': customer.name,
                    'province': customer.province,
                    'city': customer.city,
                    'address': customer.address,
                    'discard_time': datetime.fromtimestamp(customer.discard_time).strftime('%Y-%m-%d %H:%M:%S') if customer.discard_time > 0 else '',
                    'create_time': customer.create_time.strftime('%Y-%m-%d %H:%M:%S')
                })
            
            return JsonResponse({
                'code': 0,
                'msg': '',
                'count': total_count,
                'data': data
            })
        except Exception as e:
            return JsonResponse({'code': 1, 'msg': f'获取失败: {str(e)}'})

# 客户订单管理视图
class CustomerOrderListView(LoginRequiredMixin, ListView):
    """客户订单列表视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = CustomerOrder
    template_name = 'customer/order/list.html'
    context_object_name = 'orders'

    def get_queryset(self):
        # 获取当前用户有权限查看的客户ID列表
        user = self.request.user
        if hasattr(user, 'is_superuser') and user.is_superuser:
            # 超级管理员可以查看所有订单
            return CustomerOrder.objects.filter(delete_time=0)
        else:
            # 获取当前用户有权限查看的客户
            allowed_customers = Customer.objects.filter(
                models.Q(belong_uid=user.id) | 
                models.Q(share_ids__contains=str(user.id))
            ).filter(delete_time=0)
            # 只显示这些客户的订单
            return CustomerOrder.objects.filter(
                customer__in=allowed_customers,
                delete_time=0
            )

class CustomerOrderListDataView(LoginRequiredMixin, View):
    """客户订单列表数据API"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def get(self, request):
        try:
            # 获取当前用户有权限查看的客户
            user = self.request.user
            if hasattr(user, 'is_superuser') and user.is_superuser:
                # 超级管理员可以查看所有订单
                orders = CustomerOrder.objects.filter(delete_time=0).select_related('customer').order_by('-create_time')
            else:
                # 获取当前用户有权限查看的客户
                allowed_customers = Customer.objects.filter(
                    models.Q(belong_uid=user.id) | 
                    models.Q(share_ids__contains=str(user.id))
                ).filter(delete_time=0)
                # 只显示这些客户的订单
                orders = CustomerOrder.objects.filter(
                    customer__in=allowed_customers,
                    delete_time=0
                ).select_related('customer').order_by('-create_time')
            
            # 处理搜索
            search = request.GET.get('search', '')
            if search:
                orders = orders.filter(Q(order_number__icontains=search) | Q(customer__name__icontains=search))
            
            # 分页
            page = int(request.GET.get('page', 1))
            limit = int(request.GET.get('limit', 20))
            start = (page - 1) * limit
            end = start + limit
            
            total_count = orders.count()
            paginated_orders = orders[start:end]
            
            data = []
            for order in paginated_orders:
                data.append({
                    'id': order.id,
                    'order_number': order.order_number,
                    'customer_name': order.customer.name,
                    'product_name': order.product_name,
                    'amount': str(order.amount),
                    'status': order.status,
                    'status_display': order.get_status_display(),
                    'order_date': order.order_date.strftime('%Y-%m-%d') if order.order_date else '',
                    'create_time': order.create_time.strftime('%Y-%m-%d %H:%M:%S')
                })
            
            return JsonResponse({
                'code': 0,
                'msg': '',
                'count': total_count,
                'data': data
            })
        except Exception as e:
            return JsonResponse({'code': 1, 'msg': f'获取失败: {str(e)}'})

class CustomerOrderCreateView(LoginRequiredMixin, View):
    """创建客户订单视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    
    def get(self, request):
        customer_id = request.GET.get('customer_id')
        if customer_id:
            try:
                # 获取当前用户
                user = request.user
                # 检查客户是否存在且当前用户有权限查看
                customer = get_object_or_404(Customer, id=customer_id, delete_time=0)
                
                # 检查权限
                if not (hasattr(user, 'is_superuser') and user.is_superuser):
                    if customer.belong_uid != user.id and str(user.id) not in customer.share_ids.split(','):
                        return JsonResponse({'status': 'error', 'message': '您没有权限访问此客户'}, json_dumps_params={'ensure_ascii': False})
                        
                # 获取客户的合同列表
                contracts = customer.contracts.filter(delete_time=0)
                
                # 获取产品分类和产品数据
                from apps.contract.models import ProductCate, Product
                product_categories = ProductCate.objects.filter(status=1).prefetch_related('product_set')
                
                # 获取所有启用的自定义订单字段
                from apps.customer.models import OrderField
                order_fields = OrderField.objects.filter(is_active=True)
                
                # 生成默认订单编号
                import time
                default_order_number = f'ORD{int(time.time())}{customer.id}'
                
                # 初始化空的现有值字典，避免模板中访问不存在的变量
                existing_values = {}
                
                # 导入 datetime 模块获取当前时间
                from datetime import datetime
                now = datetime.now()
                
                return render(request, 'customer/customer_order_form.html', {
                    'customer': customer,
                    'contracts': contracts,
                    'product_categories': product_categories,
                    'default_order_number': default_order_number,
                    'order_fields': order_fields,
                    'existing_values': existing_values,
                    'now': now  # 传递当前时间到模板
                })
            except Customer.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': '客户不存在'}, json_dumps_params={'ensure_ascii': False})
        # 初始化空的现有值字典，避免模板中访问不存在的变量
        existing_values = {}
        
        # 导入 datetime 模块获取当前时间
        from datetime import datetime
        now = datetime.now()
        
        return render(request, 'customer/customer_order_form.html', {
            'existing_values': existing_values,
            'now': now  # 传递当前时间到模板
        })
    
    def post(self, request):
        try:
            customer_id = request.POST.get('customer_id')
            order_number = request.POST.get('order_number')
            product_name = request.POST.get('product_name')
            amount = request.POST.get('amount')
            order_date = request.POST.get('order_date')
            status = request.POST.get('status')
            contract_id = request.POST.get('contract_id')  # 新增合同关联
            description = request.POST.get('description', '')
            remark = request.POST.get('remark', '')
            
            # 检查必填字段是否为空（包括空字符串的情况）
            required_fields = [
                ('customer_id', customer_id),
                ('order_number', order_number),
                ('product_name', product_name),
                ('amount', amount),
                ('order_date', order_date),
                ('status', status)
            ]
            
            missing_fields = []
            for field_name, field_value in required_fields:
                if not field_value or str(field_value).strip() == '':
                    missing_fields.append(field_name)
            
            if missing_fields:
                return JsonResponse({'status': 'error', 'message': f'必填字段不能为空: {", ".join(missing_fields)}'})
            
            # 获取当前用户
            user = request.user
            # 检查客户是否存在且当前用户有权限查看
            customer = get_object_or_404(Customer, id=customer_id, delete_time=0)
            
            # 检查权限
            if not (hasattr(user, 'is_superuser') and user.is_superuser):
                if customer.belong_uid != user.id and str(user.id) not in customer.share_ids.split(','):
                    return JsonResponse({'status': 'error', 'message': '您没有权限访问此客户'}, json_dumps_params={'ensure_ascii': False})
            
            # 获取关联合同（如果选择了）
            contract = None
            if contract_id:
                try:
                    contract = customer.contracts.get(id=contract_id, delete_time=0)
                except:
                    pass
            
            # 创建订单
            order = CustomerOrder.objects.create(
                customer=customer,
                contract=contract,
                order_number=order_number,
                product_name=product_name,
                amount=amount,
                order_date=order_date,
                status=status,
                description=description,
                remark=remark,
                create_user=request.user
            )
            
            # 处理自定义字段值
            from apps.customer.models import OrderField
            order_fields = OrderField.objects.filter(is_active=True)
            for field in order_fields:
                field_key = f'custom_field_{field.id}'
                if field_key in request.POST:
                    # 对于复选框类型，值可能是列表
                    if field.field_type == 'checkbox':
                        values = request.POST.getlist(field_key)
                        value = ','.join(values)
                    else:
                        value = request.POST.get(field_key)
                    
                    if value:
                        # 保存自定义字段值
                        CustomerOrderCustomFieldValue.objects.create(
                            order=order,
                            field=field,
                            value=value
                        )
            
            # 需求案例3：创建订单后自动生成待签约合同、待收款、待开票、项目
            try:
                self._auto_generate_related_records(order, request.user)
            except Exception as auto_gen_error:
                # 记录错误但不影响订单创建
                logger.error(f"自动生成相关记录失败: {str(auto_gen_error)}")
            
            return JsonResponse({'status': 'success', 'message': '订单添加成功'}, json_dumps_params={'ensure_ascii': False})
            
        except Customer.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': '客户不存在'}, json_dumps_params={'ensure_ascii': False})
        except Exception as e:
            logger.error(f"创建客户订单失败: {str(e)}")
            return JsonResponse({'status': 'error', 'message': f'添加失败: {str(e)}'}, json_dumps_params={'ensure_ascii': False})
    
    def _auto_generate_related_records(self, order, user):
        """
        自动生成与订单相关的记录：待签约合同、待收款、待开票、项目
        """
        import time
        
        # 1. 生成待签约合同记录
        try:
            from apps.customer.models import CustomerContract
            # 创建待签约合同记录
            contract_record = CustomerContract.objects.create(
                customer_id=order.customer_id,
                name=f"{order.product_name}合同",
                contract_number=f"CONT-ORD-{order.order_number}-{int(time.time())}",
                amount=order.amount,
                sign_date=order.order_date if order.order_date else None,
                end_date=None,  # 待签约，未设置结束日期
                status='pending',  # 待签约状态
                create_user_id=user.id,
                auto_generated=True  # 标记为自动生成
            )
        except Exception as e:
            logger.error(f"创建合同记录失败: {e}")
        
        # 2. 生成项目记录
        try:
            from apps.project.models import Project
            # 创建项目记录
            project_record = Project.objects.create(
                name=order.product_name,
                code=f"PROJ-ORD-{order.order_number}-{int(time.time())}",
                description=f"订单项目：{order.product_name}",
                customer_id=order.customer_id,
                contract_id=0,  # 待签约，暂不关联合同
                budget=order.amount,
                status=1,  # 未开始状态
                priority=2,  # 中等优先级
                progress=0,  # 0%进度
                creator=user,
                auto_generated=True  # 标记为自动生成
            )
        except Exception as e:
            logger.error(f"创建项目记录失败: {e}")
        
        # 3. 生成待收款记录
        try:
            from apps.finance_new.models import Invoice
            # 创建发票记录（待收款）
            invoice_data = {
                'code': f"INV-ORD-{order.order_number}-{int(time.time())}",
                'customer_id': order.customer_id,
                'amount': order.amount,
                'applicant': user,
                'invoice_title': "",  # 留空，后续填写
                'invoice_status': 'draft',  # 草稿状态
                'enter_status': 0,  # 未回款状态
                'auto_generated': True  # 标记为自动生成
            }
            # 如果项目记录创建成功，添加项目关联
            if 'project_record' in locals() and project_record:
                invoice_data['project_id'] = project_record.id
            
            invoice_record = Invoice.objects.create(**invoice_data)
        except Exception as e:
            logger.error(f"创建发票记录失败: {e}")

class CustomerOrderUpdateView(LoginRequiredMixin, UpdateView):
    """编辑客户订单视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = CustomerOrder
    template_name = 'customer/order/edit.html'
    fields = ['customer', 'order_number', 'product_name', 'amount', 'order_date', 'status', 'description', 'remark']
    context_object_name = 'order'
    success_url = reverse_lazy('customer:customer_order_list')
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        # 添加自定义订单字段到上下文
        from apps.customer.models import OrderField
        context['order_fields'] = OrderField.objects.filter(is_active=True)
        # 获取现有的自定义字段值
        context['custom_field_values'] = {}
        for custom_value in self.object.custom_fields.all():
            context['custom_field_values'][custom_value.field_id] = custom_value.value
        return context
    
    def post(self, request, *args, **kwargs):
        # 获取订单对象
        self.object = self.get_object()
        
        # 使用父类的post方法处理常规字段
        response = super().post(request, *args, **kwargs)
        
        # 处理自定义字段值
        from apps.customer.models import OrderField
        order_fields = OrderField.objects.filter(is_active=True)
        
        # 删除现有的自定义字段值
        self.object.custom_fields.all().delete()
        
        # 添加新的自定义字段值
        for field in order_fields:
            field_key = f'custom_field_{field.id}'
            if field_key in request.POST:
                # 对于复选框类型，值可能是列表
                if field.field_type == 'checkbox':
                    values = request.POST.getlist(field_key)
                    value = ','.join(values)
                else:
                    value = request.POST.get(field_key)
                
                if value:
                    # 保存自定义字段值
                    CustomerOrderCustomFieldValue.objects.create(
                        order=self.object,
                        field=field,
                        value=value
                    )
        
        return response

class CustomerOrderDeleteView(LoginRequiredMixin, DeleteView):
    """删除客户订单视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = CustomerOrder
    success_url = reverse_lazy('customer:customer_order_list')


class CustomerOrderDetailView(LoginRequiredMixin, DetailView):
    """订单详情视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = CustomerOrder
    template_name = 'customer/order/detail.html'
    context_object_name = 'order'


class CustomerOrderPaymentView(LoginRequiredMixin, View):
    """订单收款视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def post(self, request, pk):
        try:
            # 获取订单对象
            order = get_object_or_404(CustomerOrder, pk=pk)
            
            # 获取请求数据
            amount = request.POST.get('amount')
            payment_method = request.POST.get('payment_method')
            payment_date = request.POST.get('payment_date')
            remark = request.POST.get('remark')
            
            # 基本验证
            if not amount:
                return JsonResponse({'status': 'error', 'message': '收款金额不能为空'})
            
            try:
                amount = float(amount)
                if amount <= 0:
                    return JsonResponse({'status': 'error', 'message': '收款金额必须大于0'})
            except ValueError:
                return JsonResponse({'status': 'error', 'message': '收款金额格式不正确'})
            
            if not payment_method:
                return JsonResponse({'status': 'error', 'message': '请选择付款方式'})
            
            if not payment_date:
                return JsonResponse({'status': 'error', 'message': '请选择收款日期'})
            
            # 使用事务处理收款操作
            with transaction.atomic():
                # 更新订单状态和收款信息
                # 根据模型中的字段进行更新
                order.finance_status = 'synced'  # 更新为已同步状态
                
                # 记录收款信息到备注中
                if remark:
                    current_remark = order.remark or ''
                    payment_info = f"\n--- 收款信息 ---\n收款金额: {amount}\n付款方式: {payment_method}\n收款日期: {payment_date}\n备注: {remark}\n--- 收款信息结束 ---".strip()
                    order.remark = current_remark + ("\n" if current_remark else "") + payment_info
                
                # 更新订单状态为已完成
                order.status = 'completed'
                order.save()
                
                # 这里可以根据实际需求添加更复杂的业务逻辑
                # 例如创建收款记录、更新财务系统等
                
            return JsonResponse({'status': 'success', 'message': '收款操作成功'})
        except Exception as e:
            logger.error(f"订单收款失败 (订单ID: {pk}): {str(e)}")
            return JsonResponse({'status': 'error', 'message': '收款操作失败，请稍后重试'})


class CustomerOrderBatchDeleteView(LoginRequiredMixin, View):
    """批量删除订单视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    
    def post(self, request):
        try:
            data = json.loads(request.body)
            ids = data.get('ids', [])
            if not ids:
                return JsonResponse({'status': 'error', 'message': '请选择要删除的订单'})
            
            CustomerOrder.objects.filter(id__in=ids).delete()
            return JsonResponse({'status': 'success', 'message': f'成功删除{len(ids)}个订单'})
        except json.JSONDecodeError:
            return JsonResponse({'status': 'error', 'message': '无效的请求数据'})
        except Exception as e:
            return JsonResponse({'status': 'error', 'message': str(e)})

# 机会线索管理视图
class OpportunityListView(LoginRequiredMixin, ListView):
    """机会线索列表视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = Customer
    template_name = 'customer/opportunity_list.html'
    context_object_name = 'opportunities'

    def get_queryset(self):
        # 机会线索：意向状态为高意向的客户
        queryset = Customer.objects.filter(delete_time=0, intent_status__in=[1, 2, 3])
        
        # 添加数据权限过滤
        user = self.request.user
        
        # 超级管理员可以查看所有客户
        if hasattr(user, 'is_superuser') and user.is_superuser:
            return queryset
        
        # 数据权限过滤：只能查看自己的客户及共享给自己的客户
        queryset = queryset.filter(
            models.Q(belong_uid=user.id) | 
            models.Q(share_ids__contains=str(user.id))
        )
        
        return queryset

class OpportunityListDataView(LoginRequiredMixin, View):
    """机会线索列表数据API"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def get(self, request):
        try:
            queryset = Customer.objects.filter(delete_time=0, intent_status__in=[1, 2, 3])
            
            # 添加数据权限过滤
            user = self.request.user
            
            # 超级管理员可以查看所有客户
            if not (hasattr(user, 'is_superuser') and user.is_superuser):
                # 数据权限过滤：只能查看自己的客户及共享给自己的客户
                queryset = queryset.filter(
                    models.Q(belong_uid=user.id) | 
                    models.Q(share_ids__contains=str(user.id))
                )
            
            # 处理搜索
            search = request.GET.get('search', '')
            if search:
                queryset = queryset.filter(name__icontains=search)
            
            # 分页
            page = int(request.GET.get('page', 1))
            limit = int(request.GET.get('limit', 20))
            start = (page - 1) * limit
            end = start + limit
            
            total_count = queryset.count()
            paginated_queryset = queryset[start:end]
            
            data = []
            for customer in paginated_queryset:
                data.append({
                    'id': customer.id,
                    'name': customer.name,
                    'province': customer.province,
                    'city': customer.city,
                    'intent_status': customer.intent_status,
                    'follow_time': customer.get_follow_time_display(),
                    'next_time': customer.get_next_time_display(),
                    'create_time': customer.create_time.strftime('%Y-%m-%d %H:%M:%S')
                })
            
            return JsonResponse({
                'code': 0,
                'msg': '',
                'count': total_count,
                'data': data
            })
        except Exception as e:
            return JsonResponse({'code': 1, 'msg': f'获取失败: {str(e)}'})

# 跟进记录管理视图
class FollowRecordListView(LoginRequiredMixin, ListView):
    """跟进记录列表视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = FollowRecord
    template_name = 'customer/follow_record_list.html'
    context_object_name = 'records'

    def get_queryset(self):
        # 获取当前用户有权限查看的客户ID列表
        user = self.request.user
        if hasattr(user, 'is_superuser') and user.is_superuser:
            # 超级管理员可以查看所有跟进记录
            return FollowRecord.objects.filter(delete_time=0)
        else:
            # 获取当前用户有权限查看的客户
            allowed_customers = Customer.objects.filter(
                models.Q(belong_uid=user.id) | 
                models.Q(share_ids__contains=str(user.id))
            ).filter(delete_time=0)
            # 只显示这些客户的跟进记录
            return FollowRecord.objects.filter(
                customer__in=allowed_customers,
                delete_time=0
            )

class FollowRecordListDataView(LoginRequiredMixin, View):
    """跟进记录列表数据API"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def get(self, request):
        try:
            # 获取当前用户
            user = self.request.user
            
            # 根据用户权限获取可查看的跟进记录
            if hasattr(user, 'is_superuser') and user.is_superuser:
                # 超级管理员可以查看所有跟进记录
                records = FollowRecord.objects.filter(delete_time=0).select_related('customer', 'follow_user').order_by('-follow_time')
            else:
                # 获取当前用户有权限查看的客户
                allowed_customers = Customer.objects.filter(
                    models.Q(belong_uid=user.id) | 
                    models.Q(share_ids__contains=str(user.id))
                ).filter(delete_time=0)
                # 只显示这些客户的跟进记录
                records = FollowRecord.objects.filter(
                    customer__in=allowed_customers,
                    delete_time=0
                ).select_related('customer', 'follow_user').order_by('-follow_time')
            
            # 处理搜索
            search = request.GET.get('search', '')
            if search:
                records = records.filter(Q(customer__name__icontains=search) | Q(content__icontains=search))
            
            # 分页
            page = int(request.GET.get('page', 1))
            limit = int(request.GET.get('limit', 20))
            start = (page - 1) * limit
            end = start + limit
            
            total_count = records.count()
            paginated_records = records[start:end]
            
            data = []
            for record in paginated_records:
                data.append({
                    'id': record.id,
                    'customer_name': record.customer.name,
                    'follow_type': record.follow_type,
                    'follow_type_display': record.get_follow_type_display(),
                    'content': record.content,
                    'follow_user': record.follow_user.username,
                    'follow_time': record.follow_time.strftime('%Y-%m-%d %H:%M:%S'),
                    'next_follow_time': record.next_follow_time.strftime('%Y-%m-%d %H:%M:%S') if record.next_follow_time else ''
                })
            
            return JsonResponse({
                'code': 0,
                'msg': '',
                'count': total_count,
                'data': data
            })
        except Exception as e:
            return JsonResponse({'code': 1, 'msg': f'获取失败: {str(e)}'})

class FollowRecordCreateView(LoginRequiredMixin, View):
    """创建跟进记录视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    
    def get(self, request):
        customer_id = request.GET.get('customer_id')
        if customer_id:
            try:
                # 获取当前用户
                user = request.user
                # 检查客户是否存在且当前用户有权限查看
                customer = get_object_or_404(Customer, id=customer_id, delete_time=0)
                
                # 检查权限
                if not (hasattr(user, 'is_superuser') and user.is_superuser):
                    if customer.belong_uid != user.id and str(user.id) not in customer.share_ids.split(','):
                        return JsonResponse({'status': 'error', 'message': '您没有权限访问此客户'}, json_dumps_params={'ensure_ascii': False})
                        
                # 获取客户的联系人列表
                contacts = customer.contacts.all()
                # 获取所有启用的自定义跟进字段
                from apps.customer.models import FollowField
                follow_fields = FollowField.objects.filter(is_active=True).order_by('sort_order')
                
                return render(request, 'customer/follow_record_form.html', {
                    'customer': customer,
                    'contacts': contacts,
                    'follow_fields': follow_fields,
                    'existing_values': {}  # 添加跟进时没有现有值
                })
            except Customer.DoesNotExist:
                return JsonResponse({'status': 'error', 'message': '客户不存在'}, json_dumps_params={'ensure_ascii': False})
        return render(request, 'customer/follow_record_form.html', {
            'existing_values': {}  # 即使没有客户ID也传递空字典
        })
    
    def post(self, request):
        try:
            customer_id = request.POST.get('customer_id')
            follow_type = request.POST.get('follow_type')
            contact_id = request.POST.get('contact_id')  # 新增联系人选择
            content = request.POST.get('content')
            next_follow_time = request.POST.get('next_follow_time')
            
            # 检查必填字段是否为空
            missing_fields = []
            if not customer_id or customer_id.strip() == '':
                missing_fields.append('客户ID')
            if not follow_type or follow_type.strip() == '':
                missing_fields.append('跟进方式')
            if not content or content.strip() == '':
                missing_fields.append('跟进内容')
            
            if missing_fields:
                return JsonResponse({'status': 'error', 'message': f'以下字段不能为空: {", ".join(missing_fields)}'})
            
            # 获取当前用户
            user = request.user
            # 检查客户是否存在且当前用户有权限查看
            customer = get_object_or_404(Customer, id=customer_id, delete_time=0)
            
            # 检查权限
            if not (hasattr(user, 'is_superuser') and user.is_superuser):
                if customer.belong_uid != user.id and str(user.id) not in customer.share_ids.split(','):
                    return JsonResponse({'status': 'error', 'message': '您没有权限访问此客户'}, json_dumps_params={'ensure_ascii': False})
            
            # 创建跟进记录
            follow_record = FollowRecord.objects.create(
                customer=customer,
                follow_type=follow_type,
                content=content,
                follow_user=request.user,
                next_follow_time=next_follow_time if next_follow_time else None
            )
            
            # 如果选择了联系人，可以在内容中记录
            if contact_id:
                try:
                    contact = customer.contacts.get(id=contact_id)
                    follow_record.content = f"联系人：{contact.contact_person}({contact.phone})\n{content}"
                    follow_record.save()
                except:
                    pass
            
            # 保存自定义字段值
            from apps.customer.models import FollowField, FollowRecordCustomFieldValue
            follow_fields = FollowField.objects.filter(is_active=True)
            
            for field in follow_fields:
                field_value = request.POST.get(f'custom_field_{field.id}')
                if field_value is not None:
                    # 处理复选框类型
                    if field.field_type == 'checkbox':
                        selected_values = request.POST.getlist(f'custom_field_{field.id}')
                        field_value = ','.join(selected_values)
                    
                    # 保存或更新自定义字段值
                    FollowRecordCustomFieldValue.objects.update_or_create(
                        follow_record=follow_record,
                        field=field,
                        defaults={'value': field_value}
                    )
            
            # 更新客户的最新跟进时间
            customer.follow_time = int(timezone.now().timestamp())
            customer.save()
            
            return JsonResponse({'status': 'success', 'message': '跟进记录添加成功'}, json_dumps_params={'ensure_ascii': False})
            
        except Customer.DoesNotExist:
            return JsonResponse({'status': 'error', 'message': '客户不存在'}, json_dumps_params={'ensure_ascii': False})
        except Exception as e:
            logger.error(f"创建跟进记录失败: {str(e)}")
            return JsonResponse({'status': 'error', 'message': f'添加失败: {str(e)}'}, json_dumps_params={'ensure_ascii': False})

class FollowRecordUpdateView(LoginRequiredMixin, UpdateView):
    """编辑跟进记录视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    model = FollowRecord
    template_name = 'customer/follow_record_form.html'
    fields = ['customer', 'follow_type', 'content', 'next_follow_time']
    success_url = reverse_lazy('customer:follow_record_list')
    
    def get_context_data(self, **kwargs):
        """获取上下文数据，添加自定义跟进字段"""
        context = super().get_context_data(**kwargs)
        # 获取所有启用的自定义跟进字段
        from customer.models import FollowField
        context['follow_fields'] = FollowField.objects.filter(is_active=True)
        
        # 获取现有的自定义字段值
        from customer.models import FollowRecordCustomFieldValue
        record = self.object
        existing_values = {}
        if record:
            custom_values = FollowRecordCustomFieldValue.objects.filter(follow_record=record)
            for cv in custom_values:
                existing_values[cv.field_id] = cv.value
        context['existing_values'] = existing_values
        
        return context
    
    def post(self, request, *args, **kwargs):
        """处理表单提交，保存自定义字段值"""
        self.object = self.get_object()
        response = super().post(request, *args, **kwargs)
        
        if self.object:
            # 处理自定义字段值
            from customer.models import FollowField, FollowRecordCustomFieldValue
            follow_fields = FollowField.objects.filter(is_active=True)
            
            # 删除旧的自定义字段值
            FollowRecordCustomFieldValue.objects.filter(follow_record=self.object).delete()
            
            # 保存新的自定义字段值
            for field in follow_fields:
                field_key = f'custom_field_{field.id}'
                
                # 处理复选框类型
                if field.field_type == 'checkbox':
                    # 获取所有选中的值
                    checkbox_values = request.POST.getlist(field_key)
                    if checkbox_values:
                        value = ','.join(checkbox_values)
                        FollowRecordCustomFieldValue.objects.create(
                            follow_record=self.object,
                            field_id=field.id,
                            value=value
                        )
                else:
                    # 处理其他类型
                    value = request.POST.get(field_key)
                    if value is not None and value.strip() != '':
                        FollowRecordCustomFieldValue.objects.create(
                            follow_record=self.object,
                            field_id=field.id,
                            value=value.strip()
                        )
        
        return response

# 拨号记录管理视图（简单实现）
class CallRecordListView(LoginRequiredMixin, TemplateView):
    """拨号记录列表视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    template_name = 'customer/call_record_list.html'

class CallRecordListDataView(LoginRequiredMixin, View):
    """拨号记录列表数据API"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def get(self, request):
        from .models import CallRecord
        from django.db.models import Count, Sum, Q
        from django.utils import timezone
        
        # 获取当前用户的拨号记录
        queryset = CallRecord.objects.filter(create_user=request.user)
        
        # 处理搜索条件
        search = request.GET.get('search', '')
        if search:
            queryset = queryset.filter(Q(phone__icontains=search) | Q(customer_name__icontains=search))
        
        # 处理分页
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 20))
        start = (page - 1) * limit
        end = start + limit
        
        # 获取总记录数
        total_count = queryset.count()
        
        # 获取分页数据
        call_records = list(queryset.order_by('-call_time')[start:end].values())
        
        # 获取所有电话号码的拨号次数
        phone_call_counts = CallRecord.objects.filter(create_user=request.user).values('phone').annotate(
            total_count=Count('id')
        )
        
        # 构建电话号码到拨号次数的映射
        call_count_map = {item['phone']: item['total_count'] for item in phone_call_counts}
        
        # 为每条记录添加累计拨号次数
        for record in call_records:
            record['call_count'] = call_count_map.get(record['phone'], 0)
            
            # 格式化通话状态
            status_map = {
                0: '未接通',
                1: '已通话',
                2: '呼叫失败',
                3: '通话中'
            }
            record['status'] = status_map.get(record['status'], '未知')
            
            # 格式化通话时长（秒转分:秒）
            duration = record['duration']
            minutes = duration // 60
            seconds = duration % 60
            record['duration'] = f'{minutes:02d}:{seconds:02d}'
            
            # 格式化拨号时间
            call_time = record['call_time']
            from datetime import datetime
            from django.utils.timezone import make_aware, get_current_timezone, is_aware
            
            try:
                if isinstance(call_time, str):
                    # 如果已经是字符串，解析并转换为上海时间
                    if call_time.endswith('Z'):
                        # UTC时间字符串
                        try:
                            # 解析带毫秒的UTC时间字符串
                            utc_time = datetime.strptime(call_time, '%Y-%m-%dT%H:%M:%S.%fZ')
                        except ValueError:
                            # 解析不带毫秒的UTC时间字符串
                            utc_time = datetime.strptime(call_time, '%Y-%m-%dT%H:%M:%SZ')
                        
                        # 转换为带时区的datetime对象
                        aware_time = make_aware(utc_time)
                        # 转换为上海时间
                        local_time = aware_time.astimezone(get_current_timezone())
                        # 格式化本地时间
                        record['call_time'] = local_time.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        # 非UTC时间字符串，尝试解析
                        try:
                            # 尝试解析为datetime对象
                            dt = datetime.strptime(call_time, '%Y-%m-%d %H:%M:%S')
                            # 格式化时间
                            record['call_time'] = dt.strftime('%Y-%m-%d %H:%M:%S')
                        except ValueError:
                            # 解析失败，直接使用
                            record['call_time'] = call_time
                else:
                    # 直接格式化时间
                    if is_aware(call_time):
                        # 带时区的datetime对象，转换为本地时间
                        local_time = call_time.astimezone(get_current_timezone())
                        record['call_time'] = local_time.strftime('%Y-%m-%d %H:%M:%S')
                    else:
                        # 不带时区的datetime对象，直接格式化
                        record['call_time'] = call_time.strftime('%Y-%m-%d %H:%M:%S')
            except Exception as e:
                # 记录错误并使用当前时间
                logger.error(f'时间格式化错误: {e}, 原始时间: {call_time}')
                # 使用当前时间作为替代
                record['call_time'] = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # 获取统计数据
        today = timezone.now().date()
        today_start = timezone.datetime(today.year, today.month, today.day, 0, 0, 0)
        
        # 今日拨号总数量
        today_call_count = CallRecord.objects.filter(
            create_user=request.user,
            call_time__gte=today_start
        ).count()
        
        # 系统累计拨号总数量
        total_call_count = CallRecord.objects.filter(create_user=request.user).count()
        
        # 今日通话总时长（秒）
        today_duration = CallRecord.objects.filter(
            create_user=request.user,
            call_time__gte=today_start,
            status=1  # 已接通
        ).aggregate(total_duration=Sum('duration'))['total_duration'] or 0
        
        # 系统累计通话总时长（秒）
        total_duration = CallRecord.objects.filter(
            create_user=request.user,
            status=1  # 已接通
        ).aggregate(total_duration=Sum('duration'))['total_duration'] or 0
        
        # 今日沟通客户总数
        today_unique_customers = CallRecord.objects.filter(
            create_user=request.user,
            call_time__gte=today_start
        ).values('customer_id').distinct().count()
        
        # 返回响应，包含统计信息和记录数据
        return JsonResponse({
            'code': 0,
            'msg': '',
            'count': total_count,
            'data': call_records,
            'statistics': {
                'today_call_count': today_call_count,
                'total_call_count': total_call_count,
                'today_duration': today_duration,
                'total_duration': total_duration,
                'today_unique_customers': today_unique_customers
            }
        })

# 批量导入相关功能
import os
import tempfile
import pandas as pd
from django.http import HttpResponse
from django.core.cache import cache
from django.conf import settings

@login_required
def download_template(request):
    """下载客户导入模板"""
    try:
        # 创建Excel模板
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "客户导入模板"
        
        # 设置表头
        headers = [
            '客户名称*', '联系人*', '联系电话*', '邮箱', '地址', 
            '省份', '城市', '区县', '客户描述', '主要经营业务', '备注信息'
        ]
        
        # 写入表头
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # 添加表头数据，不包含示例数据
        
        # 调整列宽
        for col in range(1, len(headers) + 1):
            ws.column_dimensions[openpyxl.utils.get_column_letter(col)].width = 15
        
        # 创建响应
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = 'attachment; filename="customer_import_template.xlsx"'
        
        wb.save(response)
        return response
        
    except Exception as e:
        logger.error(f'下载模板失败: {str(e)}')
        return JsonResponse({'code': 1, 'msg': f'下载模板失败: {str(e)}'})

@login_required
def upload_import_file(request):
    """上传导入文件"""
    if request.method != 'POST':
        return JsonResponse({'code': 1, 'msg': '请求方法错误'})
    
    try:
        if 'file' not in request.FILES:
            return JsonResponse({'code': 1, 'msg': '请选择要上传的文件'})
        
        file = request.FILES['file']
        
        # 验证文件类型
        if not file.name.endswith(('.xlsx', '.xls')):
            return JsonResponse({'code': 1, 'msg': '请上传Excel文件(.xlsx或.xls格式)'})
        
        # 验证文件大小
        if file.size > 10 * 1024 * 1024:  # 10MB
            return JsonResponse({'code': 1, 'msg': '文件大小不能超过10MB'})
        
        # 保存临时文件
        file_id = str(uuid.uuid4())
        temp_dir = os.path.join(settings.MEDIA_ROOT, 'temp', 'import')
        os.makedirs(temp_dir, exist_ok=True)
        
        file_path = os.path.join(temp_dir, f'{file_id}.xlsx')
        with open(file_path, 'wb') as f:
            for chunk in file.chunks():
                f.write(chunk)
        
        # 读取Excel文件获取表头
        try:
            df = pd.read_excel(file_path, nrows=0)  # 只读取表头
            headers = df.columns.tolist()
        except Exception as e:
            os.remove(file_path)
            return JsonResponse({'code': 1, 'msg': f'文件格式错误: {str(e)}'})
        
        # 缓存文件信息
        cache_key = f'import_file_{file_id}'
        cache.set(cache_key, {
            'file_path': file_path,
            'headers': headers,
            'upload_time': timezone.now().isoformat()
        }, timeout=3600)  # 1小时过期
        
        return JsonResponse({
            'code': 0,
            'msg': '文件上传成功',
            'data': {
                'file_id': file_id,
                'headers': headers
            }
        })
        
    except Exception as e:
        logger.error(f'上传文件失败: {str(e)}')
        return JsonResponse({'code': 1, 'msg': f'上传文件失败: {str(e)}'})

@login_required
def process_import(request):
    """处理导入数据"""
    if request.method != 'POST':
        return JsonResponse({'code': 1, 'msg': '请求方法错误'})
    
    try:
        file_id = request.POST.get('file_id')
        field_mapping = json.loads(request.POST.get('field_mapping', '{}'))
        
        if not file_id:
            return JsonResponse({'code': 1, 'msg': '文件ID不能为空'})
        
        # 获取文件信息
        cache_key = f'import_file_{file_id}'
        file_info = cache.get(cache_key)
        if not file_info:
            return JsonResponse({'code': 1, 'msg': '文件已过期，请重新上传'})
        
        # 创建导入任务
        task_id = str(uuid.uuid4())
        
        # 异步处理导入（这里简化为同步处理）
        result = _process_import_data(file_info['file_path'], field_mapping, request.user, task_id)
        
        return JsonResponse({
            'code': 0,
            'msg': '导入任务已创建',
            'data': {'task_id': task_id}
        })
        
    except Exception as e:
        logger.error(f'处理导入失败: {str(e)}')
        return JsonResponse({'code': 1, 'msg': f'处理导入失败: {str(e)}'})

def _process_import_data(file_path, field_mapping, user, task_id):
    """处理导入数据的核心逻辑"""
    try:
        # 设置初始进度
        progress_key = f'import_progress_{task_id}'
        cache.set(progress_key, {
            'status': 'processing',
            'percent': 0,
            'message': '开始处理数据...',
            'result': None
        }, timeout=3600)
        
        # 读取Excel文件
        df = pd.read_excel(file_path)
        total_rows = len(df)
        
        success_count = 0
        error_count = 0
        errors = []
        
        # 获取基础数据
        try:
            from apps.customer.models import CustomerSource, CustomerGrade, CustomerIntent
            default_source = CustomerSource.objects.filter(delete_time=0).first()
            default_grade = CustomerGrade.objects.filter(delete_time=0).first()
            default_intent = CustomerIntent.objects.filter(delete_time=0).first()
        except:
            default_source = None
            default_grade = None
            default_intent = None
        
        # 逐行处理数据
        for index, row in df.iterrows():
            try:
                # 更新进度
                percent = int((index + 1) / total_rows * 100)
                cache.set(progress_key, {
                    'status': 'processing',
                    'percent': percent,
                    'message': f'正在处理第 {index + 1} 行数据...',
                    'result': None
                }, timeout=3600)
                
                # 提取数据
                customer_data = {}
                for field_key, column_index in field_mapping.items():
                    if column_index < len(row):
                        value = row.iloc[column_index]
                        if pd.notna(value):
                            customer_data[field_key] = str(value).strip()
                
                # 验证必填字段
                required_fields = ['name', 'contact_person', 'phone']
                missing_fields = [field for field in required_fields if not customer_data.get(field)]
                if missing_fields:
                    errors.append({
                        'row': index + 2,  # Excel行号从2开始
                        'message': f'缺少必填字段: {", ".join(missing_fields)}'
                    })
                    error_count += 1
                    continue
                
                # 检查客户是否已存在
                existing_customer = Customer.objects.filter(
                    name=customer_data['name'],
                    delete_time=0
                ).first()
                
                if existing_customer:
                    errors.append({
                        'row': index + 2,
                        'message': f'客户 "{customer_data["name"]}" 已存在'
                    })
                    error_count += 1
                    continue
                
                # 创建客户记录
                with transaction.atomic():
                    customer = Customer.objects.create(
                        name=customer_data.get('name', ''),
                        contact_person=customer_data.get('contact_person', ''),
                        phone=customer_data.get('phone', ''),
                        email=customer_data.get('email', ''),
                        address=customer_data.get('address', ''),
                        province=customer_data.get('province', ''),
                        city=customer_data.get('city', ''),
                        district=customer_data.get('district', ''),
                        content=customer_data.get('content', ''),
                        market=customer_data.get('market', ''),
                        remark=customer_data.get('remark', ''),
                        source=default_source,
                        grade=default_grade,
                        intent=default_intent,
                        admin=user,
                        create_time=timezone.now(),
                        update_time=timezone.now()
                    )
                    
                    success_count += 1
                    
            except Exception as e:
                errors.append({
                    'row': index + 2,
                    'message': f'导入失败: {str(e)}'
                })
                error_count += 1
                continue
        
        # 设置完成状态
        result = {
            'total': total_rows,
            'success': success_count,
            'error': error_count,
            'errors': errors
        }
        
        cache.set(progress_key, {
            'status': 'completed',
            'percent': 100,
            'message': '导入完成',
            'result': result
        }, timeout=3600)
        
        # 清理临时文件
        try:
            os.remove(file_path)
        except:
            pass
        
        return result
        
    except Exception as e:
        logger.error(f'导入数据处理失败: {str(e)}')
        cache.set(progress_key, {
            'status': 'failed',
            'percent': 0,
            'message': f'导入失败: {str(e)}',
            'result': None
        }, timeout=3600)
        return None

@login_required
def import_progress(request):
    """获取导入进度"""
    task_id = request.GET.get('task_id')
    if not task_id:
        return JsonResponse({'code': 1, 'msg': '任务ID不能为空'})
    
    progress_key = f'import_progress_{task_id}'
    progress = cache.get(progress_key)
    
    if not progress:
        return JsonResponse({'code': 1, 'msg': '任务不存在或已过期'})
    
    return JsonResponse({
        'code': 0,
        'msg': '获取进度成功',
        'data': progress
    })
# AI机器人管理

class AIRobotView(LoginRequiredMixin, TemplateView):
    """AI机器人管理页面"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    template_name = 'customer/ai_robot.html'
    
    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context['title'] = 'AI机器人管理'
        return context

class AIRobotDataView(LoginRequiredMixin, View):
    """AI机器人数据API"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    
    def get(self, request):
        try:
            # 统计数据
            total_customers = Customer.objects.filter(delete_time=0).count()
            public_customers = Customer.objects.filter(delete_time=0, belong_uid=0).count()
            personal_customers = Customer.objects.filter(delete_time=0, belong_uid__gt=0).count()
            abandoned_customers = Customer.objects.filter(delete_time=0, discard_time__gt=0).count()
            
            # AI机器人功能数据
            ai_robot_data = [
                {
                    'id': 1,
                    'name': '智能客户分类机器人',
                    'status': 1,
                    'description': '自动根据客户特征进行分类，帮助销售更好地识别优质客户',
                    'active_customers': total_customers,
                    'success_rate': 0.85,
                    'last_run_time': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                },
                {
                    'id': 2,
                    'name': '智能客户画像生成机器人',
                    'status': 1,
                    'description': '基于客户数据自动生成详细的客户画像，包括客户需求、购买意向等',
                    'active_customers': personal_customers,
                    'success_rate': 0.82,
                    'last_run_time': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                },
                {
                    'id': 3,
                    'name': '智能客户跟进建议机器人',
                    'status': 1,
                    'description': '根据客户互动历史和行为数据，提供个性化的跟进建议',
                    'active_customers': personal_customers,
                    'success_rate': 0.78,
                    'last_run_time': timezone.now().strftime('%Y-%m-%d %H:%M:%S')
                },
                {
                    'id': 4,
                    'name': '智能客户分配机器人',
                    'status': 0,
                    'description': '根据销售的工作负载和专长，自动分配公海客户',
                    'active_customers': public_customers,
                    'success_rate': 0.80,
                    'last_run_time': (timezone.now() - timedelta(days=1)).strftime('%Y-%m-%d %H:%M:%S')
                }
            ]
            
            # 准备响应数据
            data = {
                'code': 0,
                'msg': '获取数据成功',
                'count': len(ai_robot_data),
                'data': ai_robot_data,
                'statistics': {
                    'total_customers': total_customers,
                    'public_customers': public_customers,
                    'personal_customers': personal_customers,
                    'abandoned_customers': abandoned_customers
                }
            }
            return JsonResponse(data)
        except Exception as e:
            logger.error(f'获取AI机器人数据失败: {str(e)}')
            return JsonResponse({'code': 1, 'msg': f'获取数据失败: {str(e)}'})

# 爬虫任务批量删除
class SpiderTaskBatchDeleteView(LoginRequiredMixin, View):
    """爬虫任务批量删除"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    
    def post(self, request):
        try:
            import json
            data = json.loads(request.body)
            ids = data.get('ids', [])
            
            if not ids:
                return JsonResponse({'status': 'error', 'message': '请选择要删除的任务'}, status=400)
            
            # 这里可以添加实际的删除逻辑
            # SpiderTask.objects.filter(id__in=ids).delete()
            
            return JsonResponse({'status': 'success', 'message': f'成功删除{len(ids)}个任务'})
        except Exception as e:
            logger.error(f'批量删除爬虫任务失败: {str(e)}')
            return JsonResponse({'status': 'error', 'message': f'删除失败: {str(e)}'}, status=500)

# 爬虫任务启动和停止
class SpiderTaskStartView(LoginRequiredMixin, View):
    """启动爬虫任务"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    
    def post(self, request, pk):
        try:
            # 这里可以添加实际的启动逻辑
            # task = SpiderTask.objects.get(id=pk)
            # task.start()
            
            return JsonResponse({'status': 'success', 'message': '任务启动成功'})
        except Exception as e:
            logger.error(f'启动爬虫任务失败: {str(e)}')
            return JsonResponse({'status': 'error', 'message': f'启动失败: {str(e)}'}, status=500)

class SpiderTaskStopView(LoginRequiredMixin, View):
    """停止爬虫任务"""
    login_url = '/user/login/'
    redirect_field_name = 'next'
    
    def post(self, request, pk):
        try:
            # 这里可以添加实际的停止逻辑
            # task = SpiderTask.objects.get(id=pk)
            # task.stop()
            
            return JsonResponse({'status': 'success', 'message': '任务停止成功'})
        except Exception as e:
            logger.error(f'停止爬虫任务失败: {str(e)}')
            return JsonResponse({'status': 'error', 'message': f'停止失败: {str(e)}'}, status=500)




# 客户字段管理视图
@login_required
def customer_field_page(request):
    """客户字段页面"""
    return render(request, 'customer/customer_field_list.html')

@login_required
def customer_field_list(request):
    """客户字段列表"""
    # 视图实现已迁移到下方


@login_required
def customer_field_form(request, pk=None):
    """客户字段表单"""
    # 视图实现已迁移到下方


@login_required
def customer_field_list_data(request):
    """客户字段列表数据API"""
    # 视图实现已迁移到下方


@login_required
def customer_field_toggle(request, pk):
    """切换客户字段状态"""
    # 视图实现已迁移到下方


@login_required
def customer_field_delete(request, pk):
    """删除客户字段"""
    # 视图实现已迁移到下方


# 客户来源管理视图
@login_required
def customer_source_list(request):
    """客户来源列表"""
    # 视图实现已迁移到下方


@login_required
def customer_source_form(request, pk=None):
    """客户来源表单"""
    # 视图实现已迁移到下方


@login_required
def customer_source_list_data(request):
    """客户来源列表数据API"""
    # 视图实现已迁移到下方


@login_required
def customer_source_toggle(request, pk):
    """切换客户来源状态"""
    # 视图实现已迁移到下方


@login_required
def customer_source_delete(request, pk):
    """删除客户来源"""
    # 视图实现已迁移到下方


# 客户等级管理视图
@login_required
def customer_grade_list(request):
    """客户等级列表"""
    # 视图实现已迁移到下方


@login_required
def customer_grade_form(request, pk=None):
    """客户等级表单"""
    # 视图实现已迁移到下方


@login_required
def customer_grade_list_data(request):
    """客户等级列表数据API"""
    # 视图实现已迁移到下方


@login_required
def customer_grade_toggle(request, pk):
    """切换客户等级状态"""
    # 视图实现已迁移到下方


@login_required
def customer_grade_delete(request, pk):
    """删除客户等级"""
    # 视图实现已迁移到下方


# 客户意向管理视图
@login_required
def customer_intent_list(request):
    """客户意向列表"""
    # 视图实现已迁移到下方


@login_required
def customer_intent_form(request, pk=None):
    """客户意向表单"""
    # 视图实现已迁移到下方


@login_required
def customer_intent_list_data(request):
    """客户意向列表数据API"""
    # 视图实现已迁移到下方


@login_required
def customer_intent_toggle(request, pk):
    """切换客户意向状态"""
    # 视图实现已迁移到下方


@login_required
def customer_intent_delete(request, pk):
    """删除客户意向"""
    # 视图实现已迁移到下方


@login_required
def customer_intent_update_sort(request, pk):
    """更新单个客户意向的排序"""
    # 视图实现已迁移到下方


@login_required
def customer_intent_batch_update_sort(request):
    """批量更新客户意向的排序"""
    # 视图实现已迁移到下方


@login_required
@require_POST
def sip_call(request):
    """SIP拨号接口"""
    import requests
    import time
    import json
    from django.http import JsonResponse
    from .models import CallRecord
    from apps.user.models.admin import Admin
    
    try:
        # 获取请求参数
        phone = request.POST.get('phone')
        customer_id = request.POST.get('customer_id')
        customer_name = request.POST.get('customer_name', '')
        
        # 验证参数
        if not phone:
            return JsonResponse({'code': 1, 'msg': '电话号码不能为空'})
        
        # 从当前用户的Admin模型中获取SIP账号信息
        # request.user已经是Admin模型的实例，因为Django的认证系统已经配置为使用Admin模型
        sip_account = request.user.sip_account
        sip_password = request.user.sip_password
        
        # 验证SIP账号和密码是否存在
        if not sip_account or not sip_password:
            return JsonResponse({'code': 1, 'msg': '当前用户未配置SIP账号信息，请联系管理员'})
        
        # 调用LYCC系统的SIP拨号接口
        lycc_url = "http://192.168.1.200:9078"
        params = {
            'op': 'callout',
            'Exten': sip_account,
            'phone': phone,
            'flowid': f"{request.user.id}_{int(time.time() * 1000)}"
        }
        
        response = requests.get(lycc_url, params=params)
        result = response.text.strip()
        
        # 解析LYCC系统的返回结果
        if result == '100':
            # 呼叫成功，记录拨号记录
            call_record = CallRecord.objects.create(
                create_user=request.user,
                customer_id=customer_id if customer_id else None,
                customer_name=customer_name,
                phone=phone,
                status=3,  # 3表示通话中
                flow_id=params['flowid']  # 保存flow_id
            )
            return JsonResponse({'code': 0, 'msg': '呼叫成功', 'call_id': call_record.id})
        else:
            # 呼叫失败，记录失败原因
            error_map = {
                '101': '分机号不存在',
                '102': '没有空闲的线路',
                '103': '参数错误',
                '104': '外呼失败',
                '105': '分机未注册',
                '500': '其他错误'
            }
            
            error_msg = error_map.get(result, f'未知错误: {result}')
            
            CallRecord.objects.create(
                create_user=request.user,
                customer_id=customer_id if customer_id else None,
                customer_name=customer_name,
                phone=phone,
                status=2,  # 2表示拨号失败
                flow_id=params['flowid']  # 保存flow_id
            )
            return JsonResponse({'code': 1, 'msg': f'呼叫失败: {error_msg} (错误码: {result})', 'result_code': result, 'error_message': error_msg})
            
    except Exception as e:
        # 记录异常信息
        try:
            CallRecord.objects.create(
                create_user=request.user,
                customer_id=customer_id if customer_id else None,
                customer_name=customer_name,
                phone=phone,
                status=2,  # 2表示拨号失败
                flow_id=params.get('flowid', '')  # 保存flow_id
            )
        except Exception as create_error:
            logger.error(f'创建拨号记录失败: {str(create_error)}')
        return JsonResponse({'code': 1, 'msg': f'呼叫失败: {str(e)}'})

@login_required
def update_call_status(request):
    """更新通话状态"""
    import requests
    import json
    import datetime
    from django.http import JsonResponse
    from django.utils import timezone
    from .models import CallRecord
    
    try:
        # 从当前用户获取所有通话记录
        all_calls = CallRecord.objects.filter(
            create_user=request.user
        )
        
        # 调用LYCC系统的外呼记录接口，获取最新的通话状态
        lycc_url = "http://192.168.1.200:9078"
        params = {
            'op': 'outlist',
            'WorkerID': request.user.sip_account  # 使用SIP账号作为员工工号
        }
        
        response = requests.get(lycc_url, params=params)
        call_records_data = []
        
        # 尝试解析响应
        if response.text:
            try:
                # 预处理响应：去除前后空白字符，然后转义反斜杠
                raw_text = response.text.strip()
                # 替换所有反斜杠为双反斜杠以符合JSON规范
                fixed_text = raw_text.replace('\\', '\\\\')
                call_records_data = json.loads(fixed_text)
            except json.JSONDecodeError:
                # 如果解析失败，记录错误（只记录前100字符避免日志过大）
                logger.error(f'LYCC接口返回数据格式错误: {response.text[:100]}...')
        
        # 更新本地记录的状态
        updated_count = 0
        
        # 1. 先处理所有通话记录，尝试从LYCC获取最新状态
        if call_records_data:
            # 遍历所有本地记录
            for local_call in all_calls:
                # 查找对应的远程记录
                for remote_call in call_records_data:
                    # 根据flow_id或电话号码匹配（使用CallerID和CalleeID字段）
                    if (remote_call.get('FlowNo') == local_call.flow_id or 
                        remote_call.get('CallerID') == local_call.phone or 
                        remote_call.get('CalleeID') == local_call.phone):
                        
                        # 更新本地状态和时长
                        # 完整的状态映射，包括数值结果（根据文档第237行）
                        status_map = {
                            # 字符串状态
                            '呼叫中': 3,      # 3表示通话中
                            '通话中': 3,
                            '未接通': 0,      # 0表示未接通
                            '已通话': 1,      # 1表示已通话
                            '通话完成': 1,    # 通话完成对应已通话
                            '呼叫失败': 2,    # 2表示呼叫失败
                            '通话结束': 1,    # 通话结束对应已通话
                            # 数值状态（根据文档第237行）
                            0: 0,    # 0呼叫排队中 → 未接通
                            1: 3,    # 1呼叫中 → 通话中
                            3: 1,    # 3呼叫成功 → 已通话
                            4: 2,    # 4呼叫失败 → 呼叫失败
                            6: 1,    # 6已转人工 → 已通话
                            7: 2,    # 7转人工失败 → 呼叫失败
                            8: 3     # 8转人工呼叫中 → 通话中
                        }
                        
                        # 获取远程状态，根据文档可能的字段名
                        remote_status = remote_call.get('Status', '') or remote_call.get('Result', '')
                        
                        # 尝试将状态转换为数值，处理字符串状态和数值状态
                        try:
                            # 尝试转换为整数
                            remote_status = int(remote_status)
                        except (ValueError, TypeError):
                            # 如果转换失败，保持为字符串
                            pass
                        
                        # 获取新状态
                        new_status = status_map.get(remote_status, local_call.status)
                        
                        # 获取通话时长，根据文档可能的字段名
                        duration = remote_call.get('Hold', 0) or remote_call.get('Duration', 0)
                        
                        # 确保时长是整数
                        try:
                            duration = int(duration)
                        except (ValueError, TypeError):
                            duration = 0
                        
                        # 计算实际通话时长
                        if local_call.status == 3 and new_status != 3:  # 通话结束
                            # 计算从拨号时间到现在的时间差
                            from django.utils import timezone
                            now = timezone.now()
                            call_duration = int((now - local_call.call_time).total_seconds())
                            # 确保时长大于0
                            if call_duration > 0:
                                duration = call_duration
                        
                        # 只有当状态或时长有变化时才更新
                        if local_call.status != new_status or local_call.duration != duration:
                            local_call.status = new_status
                            local_call.duration = duration
                            local_call.save()
                            updated_count += 1
                        break
        
        # 2. 处理超时的"通话中"记录
        timeout_threshold = timezone.now() - datetime.timedelta(minutes=1)  # 1分钟超时，快速更新
        
        timed_out_calls = CallRecord.objects.filter(
            create_user=request.user,
            status=3,  # 3表示通话中
            call_time__lt=timeout_threshold
        )
        
        for call in timed_out_calls:
            # 超时的通话中记录标记为已通话（假设通话已完成）
            call.status = 1  # 1表示已通话
            # 设置合理的通话时长，默认为60秒
            if call.duration == 0:
                call.duration = 60
            call.save()
            updated_count += 1
        
        return JsonResponse({'code': 0, 'msg': f'成功更新{updated_count}条通话状态', 'updated_count': updated_count})
    except Exception as e:
        logger.error(f'更新通话状态失败: {str(e)}')
        return JsonResponse({'code': 1, 'msg': f'更新通话状态失败: {str(e)}'})

# 跟进字段管理视图
@login_required
def follow_field_list(request):
    """跟进字段列表"""
    from django.shortcuts import render
    
    # 准备上下文，使用静态URL路径
    context = {
        'page_title': '跟进字段管理',
        'list_url': '/customer/follow/field/list/data/',
        'add_url': '/customer/follow/field/form/',
        'edit_url': '/customer/follow/field/form/{id}/',
        'delete_url': '/customer/follow/field/delete/{id}/'
    }
    
    return render(request, 'customer/follow_field_list.html', context)

@login_required
def follow_field_form(request, pk=None):
    """跟进字段表单"""
    from apps.common.views_utils import generic_form_view
    from .models import FollowField
    from .forms import FollowFieldForm
    return generic_form_view(
        request,
        FollowField,
        FollowFieldForm,
        'customer/follow_field_form.html',
        'customer:follow_field_list',
        pk
    )

@login_required
def follow_field_list_data(request):
    """跟进字段列表数据API"""
    from .models import FollowField
    from django.core.paginator import Paginator
    from django.db.models import Q
    from django.http import JsonResponse
    
    try:
        search = request.GET.get('search', '')
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 20))
        
        # 构建查询
        query = Q()
        if search:
            query |= Q(name__icontains=search) | Q(field_name__icontains=search)
        
        # 获取数据
        fields = FollowField.objects.filter(query).order_by('sort_order', 'id')
        
        # 分页
        paginator = Paginator(fields, limit)
        page_obj = paginator.get_page(page)
        
        # 构建返回数据
        data = []
        for field in page_obj:
            data.append({
                'id': field.id,
                'name': field.name,
                'field_name': field.field_name,
                'field_type': field.field_type,
                'get_field_type_display': field.get_field_type_display(),
                'is_required': field.is_required,
                'sort_order': field.sort_order,
                'is_active': field.is_active,
                'created_at': field.created_at.strftime('%Y-%m-%d %H:%M:%S') if field.created_at else ''
            })
        
        return JsonResponse({
            'code': 0,
            'msg': '',
            'count': paginator.count,
            'data': data
        })
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'获取失败: {str(e)}'})

@login_required
def follow_field_toggle(request, pk):
    """切换跟进字段状态"""
    try:
        from .models import FollowField
        field = FollowField.objects.get(id=pk)
        field.is_active = not field.is_active
        field.save()
        
        status_text = '启用' if field.is_active else '禁用'
        return JsonResponse({'code': 0, 'msg': f'字段已{status_text}'})
        
    except FollowField.DoesNotExist:
        return JsonResponse({'code': 1, 'msg': '字段不存在'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'操作失败: {str(e)}'})

@login_required
def follow_field_delete(request, pk):
    """删除跟进字段"""
    try:
        from .models import FollowField
        field = FollowField.objects.get(id=pk)
        field.delete()
        
        return JsonResponse({'code': 0, 'msg': '删除成功'})
        
    except FollowField.DoesNotExist:
        return JsonResponse({'code': 1, 'msg': '字段不存在'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'删除失败: {str(e)}'})

# 订单字段管理视图
@login_required
def order_field_list_data(request):
    """订单字段列表数据API"""
    from .models import OrderField
    from django.core.paginator import Paginator
    from django.db.models import Q
    from django.http import JsonResponse
    
    try:
        search = request.GET.get('search', '')
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 20))
        
        # 构建查询
        query = Q()
        if search:
            query |= Q(name__icontains=search) | Q(field_name__icontains=search)
        
        # 获取数据
        fields = OrderField.objects.filter(query).order_by('sort_order', 'id')
        
        # 分页
        paginator = Paginator(fields, limit)
        page_obj = paginator.get_page(page)
        
        # 构建返回数据
        data = []
        for field in page_obj:
            data.append({
                'id': field.id,
                'name': field.name,
                'field_name': field.field_name,
                'field_type': field.field_type,
                'get_field_type_display': field.get_field_type_display(),
                'is_required': field.is_required,
                'is_summary': field.is_summary,
                'sort_order': field.sort_order,
                'is_active': field.is_active,
                'created_at': field.created_at.strftime('%Y-%m-%d %H:%M:%S') if field.created_at else ''
            })
        
        return JsonResponse({
            'code': 0,
            'msg': '',
            'count': paginator.count,
            'data': data
        })
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'获取失败: {str(e)}'})

@login_required
def order_field_toggle(request, pk):
    """切换订单字段状态"""
    try:
        from .models import OrderField
        field = OrderField.objects.get(id=pk)
        field.is_active = not field.is_active
        field.save()
        
        status_text = '启用' if field.is_active else '禁用'
        return JsonResponse({'code': 0, 'msg': f'字段已{status_text}'})
        
    except OrderField.DoesNotExist:
        return JsonResponse({'code': 1, 'msg': '字段不存在'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'操作失败: {str(e)}'})

@login_required
def order_field_delete(request, pk):
    """删除订单字段"""
    try:
        from .models import OrderField
        field = OrderField.objects.get(id=pk)
        field.delete()
        
        return JsonResponse({'code': 0, 'msg': '删除成功'})
        
    except OrderField.DoesNotExist:
        return JsonResponse({'code': 1, 'msg': '字段不存在'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'删除失败: {str(e)}'})

@login_required
def order_field_list(request):
    """订单字段列表"""
    from django.shortcuts import render
    
    # 准备上下文，使用静态URL路径
    context = {
        'page_title': '订单字段管理',
        'list_url': '/customer/order/field/list/data/',
        'add_url': '/customer/order/field/form/',
        'edit_url': '/customer/order/field/form/{id}/',
        'delete_url': '/customer/order/field/delete/{id}/'
    }
    
    return render(request, 'customer/order_field_list.html', context)

@login_required
def order_field_form(request, pk=None):
    """订单字段表单"""
    from apps.common.views_utils import generic_form_view
    from .models import OrderField
    from .forms import OrderFieldForm
    return generic_form_view(
        request,
        OrderField,
        OrderFieldForm,
        'customer/order_field_form.html',
        'customer:order_field_list',
        pk
    )


@login_required
def follow_field_sync(request):
    """同步跟进字段配置"""
    from django.http import JsonResponse
    from .models import FollowField, FollowRecord, FollowRecordCustomFieldValue
    
    if request.method == 'POST':
        try:
            # 获取所有启用的跟进字段
            active_fields = FollowField.objects.filter(is_active=True)
            
            # 获取所有跟进记录
            follow_records = FollowRecord.objects.all()
            
            # 为每个跟进记录同步字段
            sync_count = 0
            for follow_record in follow_records:
                for field in active_fields:
                    # 检查是否已存在对应的自定义字段值记录
                    custom_field_value, created = FollowRecordCustomFieldValue.objects.get_or_create(
                        follow_record=follow_record,
                        field=field,
                        defaults={'value': ''}
                    )
                    if created:
                        sync_count += 1

            # 返回成功响应
            return JsonResponse({
                'status': 'success',
                'message': f'成功同步 {sync_count} 个跟进字段配置'
            })
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'同步失败: {str(e)}'
            })
    else:
        return JsonResponse({
            'status': 'error',
            'message': '无效的请求方法'
        })


@login_required
def order_field_sync(request):
    """同步订单字段配置"""
    from django.http import JsonResponse
    from .models import OrderField, CustomerOrder, CustomerOrderCustomFieldValue
    
    if request.method == 'POST':
        try:
            # 获取所有启用的订单字段
            active_fields = OrderField.objects.filter(is_active=True)
            
            # 获取所有订单记录
            orders = CustomerOrder.objects.all()
            
            # 为每个订单同步字段
            sync_count = 0
            for order in orders:
                for field in active_fields:
                    # 检查是否已存在对应的自定义字段值记录
                    custom_field_value, created = CustomerOrderCustomFieldValue.objects.get_or_create(
                        order=order,
                        field=field,
                        defaults={'value': ''}
                    )
                    if created:
                        sync_count += 1

            # 返回成功响应
            return JsonResponse({
                'status': 'success',
                'message': f'成功同步 {sync_count} 个订单字段配置'
            })
        except Exception as e:
            return JsonResponse({
                'status': 'error',
                'message': f'同步失败: {str(e)}'
            })
    else:
        return JsonResponse({
            'status': 'error',
            'message': '无效的请求方法'
        })


class ContractAddView(LoginRequiredMixin, View):
    """合同添加页面视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def get(self, request):
        """渲染合同添加表单页面"""
        customer_id = request.GET.get('customer_id')
        
        # 生成默认合同编号（日期+序号格式）
        from django.utils import timezone
        date_str = timezone.now().strftime('%Y%m%d')
        
        # 查找当日已存在的最大序号
        last_contract = CustomerContract.objects.filter(
            contract_number__startswith=date_str
        ).order_by('-contract_number').first()
        
        if last_contract and last_contract.contract_number:
            try:
                if len(last_contract.contract_number) >= 12:
                    last_num = int(last_contract.contract_number[8:12])
                    new_num = last_num + 1
                else:
                    new_num = 1
            except (ValueError, IndexError):
                new_num = 1
        else:
            new_num = 1
        
        default_contract_number = f'{date_str}{new_num:04d}'
        
        # 加载合同分类数据
        try:
            from apps.contract.models import ContractCategory
            contract_categories = ContractCategory.objects.filter(delete_time=0).order_by('name')
        except ImportError:
            # 如果basedata模块不可用，创建一个空列表
            contract_categories = []
        
        if customer_id:
            try:
                customer = Customer.objects.get(id=customer_id, delete_time=0)
                return render(request, 'customer/add_contract.html', {
                    'customer': customer,
                    'default_contract_number': default_contract_number,
                    'contract_categories': contract_categories
                })
            except Customer.DoesNotExist:
                return render(request, 'customer/add_contract.html', {
                    'error': '客户不存在',
                    'default_contract_number': default_contract_number,
                    'contract_categories': contract_categories
                })
        
        return render(request, 'customer/add_contract.html', {
            'default_contract_number': default_contract_number,
            'contract_categories': contract_categories
        })

    def post(self, request):
        """处理合同添加表单提交"""
        from django.http import JsonResponse
        from django.forms.models import model_to_dict
        import json
        
        try:
            # 解析JSON数据
            data = json.loads(request.body)
            
            # 获取客户ID
            customer_id = data.get('customer_id')
            if not customer_id:
                return JsonResponse({
                    'code': 1,
                    'msg': '客户ID不能为空'
                })
            
            # 验证客户是否存在
            try:
                customer = Customer.objects.get(id=customer_id, delete_time=0)
            except Customer.DoesNotExist:
                return JsonResponse({
                    'code': 1,
                    'msg': '客户不存在'
                })
            
            # 验证必填字段
            required_fields = ['contract_number', 'name', 'amount', 'sign_date', 'status', 'category_id']
            for field in required_fields:
                if not data.get(field):
                    return JsonResponse({
                        'code': 1,
                        'msg': f'{field}字段不能为空'
                    })
            
            # 验证和转换日期格式
            from django.utils import timezone
            
            # 处理签订日期
            sign_date = None
            if data['sign_date']:
                try:
                    # 尝试解析日期格式
                    if isinstance(data['sign_date'], str):
                        # 如果是字符串，尝试解析为日期对象
                        sign_date = timezone.datetime.strptime(data['sign_date'], '%Y-%m-%d').date()
                    else:
                        # 如果不是字符串，直接使用
                        sign_date = data['sign_date']
                except (ValueError, TypeError):
                    return JsonResponse({
                        'code': 1,
                        'msg': '签订日期格式错误，应为YYYY-MM-DD格式'
                    })
            
            # 处理到期日期
            end_date = None
            if data.get('end_date'):
                try:
                    if isinstance(data['end_date'], str):
                        end_date = timezone.datetime.strptime(data['end_date'], '%Y-%m-%d').date()
                    else:
                        end_date = data['end_date']
                except (ValueError, TypeError):
                    return JsonResponse({
                        'code': 1,
                        'msg': '到期日期格式错误，应为YYYY-MM-DD格式'
                    })
            
            # 创建合同记录
            contract = CustomerContract.objects.create(
                customer=customer,
                contract_number=data['contract_number'],
                name=data['name'],
                amount=data['amount'],
                sign_date=sign_date,
                end_date=end_date,
                status=data['status'],
                category_id=data['category_id'],
                contract_type=data.get('contract_type', ''),
                description=data.get('description', ''),
                remark=data.get('remark', ''),
                create_user=request.user
            )
            
            # 返回成功响应
            return JsonResponse({
                'code': 0,
                'msg': '合同添加成功',
                'data': {
                    'id': contract.id,
                    'contract_number': contract.contract_number,
                    'name': contract.name
                }
            })
            
        except json.JSONDecodeError:
            return JsonResponse({
                'code': 1,
                'msg': '请求数据格式错误'
            })
        except Exception as e:
            logger.error(f"合同添加失败: {str(e)}")
            return JsonResponse({
                'code': 1,
                'msg': f'合同添加失败: {str(e)}'
            })


class CustomerSelectView(LoginRequiredMixin, View):
    """客户选择视图"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def get(self, request):
        """渲染客户选择页面"""
        return render(request, 'customer/customer_select.html')

    def post(self, request):
        """获取客户列表数据"""
        try:
            # 获取查询参数
            search = request.POST.get('search', '')
            page = int(request.POST.get('page', 1))
            limit = int(request.POST.get('limit', 10))
            
            # 构建查询条件
            queryset = Customer.objects.filter(delete_time=0)
            
            if search:
                queryset = queryset.filter(
                    Q(name__icontains=search) | 
                    Q(contacts__contact_person__icontains=search) |
                    Q(contacts__phone__icontains=search)
                ).distinct()
            
            # 分页处理
            total_count = queryset.count()
            start = (page - 1) * limit
            end = start + limit
            customers = queryset[start:end]
            
            # 构建返回数据
            data = []
            for customer in customers:
                # 获取主要联系人信息
                primary_contact = customer.contacts.filter(is_primary=True).first()
                if not primary_contact:
                    primary_contact = customer.contacts.first()
                
                contact_name = primary_contact.contact_person if primary_contact else ''
                phone = primary_contact.phone if primary_contact else ''
                
                data.append({
                    'id': customer.id,
                    'name': customer.name,
                    'contact_name': contact_name,
                    'phone': phone,
                    'address': customer.address,
                    'create_time': customer.create_time.strftime('%Y-%m-%d') if customer.create_time else ''
                })
            
            return JsonResponse({
                'code': 0,
                'msg': '',
                'count': total_count,
                'data': data
            })
            
        except Exception as e:
            logger.error(f"获取客户列表失败: {str(e)}")
            return JsonResponse({
                'code': 1,
                'msg': f'获取客户列表失败: {str(e)}'
            })


class ContractNumberCheckView(LoginRequiredMixin, View):
    """检查合同编号是否已存在"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def post(self, request):
        """检查合同编号是否已存在"""
        contract_number = request.POST.get('contract_number')
        
        if not contract_number:
            return JsonResponse({
                'exists': False,
                'msg': '合同编号不能为空'
            })
        
        # 检查编号是否已存在
        exists = CustomerContract.objects.filter(
            contract_number=contract_number
        ).exists()
        
        return JsonResponse({
            'exists': exists,
            'msg': '编号已存在' if exists else '编号可用'
        })


class ContractNumberGenerateView(LoginRequiredMixin, View):
    """生成新的合同编号"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def get(self, request):
        """生成新的合同编号"""
        from django.utils import timezone
        
        # 生成默认合同编号（日期+序号格式）
        date_str = timezone.now().strftime('%Y%m%d')
        
        # 查找当日已存在的最大序号
        last_contract = CustomerContract.objects.filter(
            contract_number__startswith=date_str
        ).order_by('-contract_number').first()
        
        if last_contract and last_contract.contract_number:
            try:
                if len(last_contract.contract_number) >= 12:
                    last_num = int(last_contract.contract_number[8:12])
                    new_num = last_num + 1
                else:
                    new_num = 1
            except (ValueError, IndexError):
                new_num = 1
        else:
            new_num = 1
        
        contract_number = f'{date_str}{new_num:04d}'
        
        return JsonResponse({
            'code': 0,
            'contract_number': contract_number,
            'msg': '合同编号生成成功'
        })


class CustomerContractListView(LoginRequiredMixin, View):
    """客户合同列表视图（用于局部刷新）"""
    login_url = '/user/login/'
    redirect_field_name = 'next'

    def get(self, request, customer_id):
        """获取客户合同列表HTML片段"""
        try:
            # 获取客户信息
            customer = Customer.objects.get(id=customer_id, delete_time=0)
            
            # 获取客户的最近合同（最多5个）
            contracts = customer.contracts.filter(delete_time=0).order_by('-create_time')[:5]
            
            # 渲染合同列表HTML片段
            return render(request, 'customer/_contract_list.html', {
                'customer': customer,
                'contracts': contracts
            })
            
        except Customer.DoesNotExist:
            return HttpResponse('<p style="text-align: center; color: #999; padding: 20px;">客户不存在</p>')
        except Exception as e:
            logger.error(f"获取客户合同列表失败: {str(e)}")
            return HttpResponse('<p style="text-align: center; color: #999; padding: 20px;">获取合同列表失败</p>')


from django.shortcuts import render, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.http import JsonResponse
from django.db.models import Q
from django.utils import timezone
from django.core.paginator import Paginator
from decimal import Decimal
from datetime import datetime
import json

from .models import CustomerSource, CustomerGrade, CustomerIntent, FollowField, OrderField
from .forms import CustomerSourceForm, CustomerGradeForm, CustomerIntentForm, FollowFieldForm, OrderFieldForm
from apps.common.views_utils import generic_list_view, generic_form_view


@login_required
def customer_source_list(request):
    return generic_list_view(
        request,
        CustomerSource,
        'customer/customer_source_list.html',
        search_fields=['title']
    )


@login_required
def customer_source_form(request, pk=None):
    return generic_form_view(
        request,
        CustomerSource,
        CustomerSourceForm,
        'customer/customer_source_form.html',
        'customer:customer_source_list',
        pk
    )


@login_required
def customer_source_list_data(request):
    try:
        search = request.GET.get('search', '')
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 20))
        
        queryset = CustomerSource.objects.filter(delete_time=0)
        
        if search:
            queryset = queryset.filter(title__icontains=search)
        
        queryset = queryset.order_by('sort', 'id')
        
        paginator = Paginator(queryset, limit)
        page_obj = paginator.get_page(page)
        
        data = []
        for obj in page_obj:
            data.append({
                'id': obj.id,
                'title': obj.title,
                'sort': obj.sort,
                'status': obj.status,
                'status_display': '启用' if obj.status == 1 else '禁用',
                'create_time': obj.create_time.strftime('%Y-%m-%d %H:%M:%S') if obj.create_time else ''
            })
        
        return JsonResponse({
            'code': 0,
            'msg': '',
            'count': paginator.count,
            'data': data
        })
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'获取失败: {str(e)}'})


@login_required
def customer_source_toggle(request, pk):
    try:
        obj = get_object_or_404(CustomerSource, pk=pk, delete_time=0)
        obj.status = 1 - obj.status
        obj.save()
        
        status_text = '启用' if obj.status == 1 else '禁用'
        return JsonResponse({'code': 0, 'msg': f'客户来源已{status_text}'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'操作失败: {str(e)}'})


@login_required
def customer_source_delete(request, pk):
    try:
        obj = get_object_or_404(CustomerSource, pk=pk, delete_time=0)
        obj.delete_time = int(timezone.now().timestamp())
        obj.save()
        
        return JsonResponse({'code': 0, 'msg': '删除成功'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'删除失败: {str(e)}'})


@login_required
def customer_grade_list(request):
    return generic_list_view(
        request,
        CustomerGrade,
        'customer/customer_grade_list.html',
        search_fields=['title']
    )


@login_required
def customer_grade_form(request, pk=None):
    return generic_form_view(
        request,
        CustomerGrade,
        CustomerGradeForm,
        'customer/customer_grade_form.html',
        'basedata:customer_grade_list',
        pk
    )


@login_required
def customer_grade_list_data(request):
    try:
        search = request.GET.get('search', '')
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 20))
        
        queryset = CustomerGrade.objects.filter(delete_time=0)
        
        if search:
            queryset = queryset.filter(title__icontains=search)
        
        queryset = queryset.order_by('sort', 'id')
        
        paginator = Paginator(queryset, limit)
        page_obj = paginator.get_page(page)
        
        data = []
        for obj in page_obj:
            data.append({
                'id': obj.id,
                'title': obj.title,
                'sort': obj.sort,
                'status': obj.status,
                'status_display': '启用' if obj.status == 1 else '禁用',
                'create_time': obj.create_time.strftime('%Y-%m-%d %H:%M:%S') if obj.create_time else ''
            })
        
        return JsonResponse({
            'code': 0,
            'msg': '',
            'count': paginator.count,
            'data': data
        })
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'获取失败: {str(e)}'})


@login_required
def customer_grade_toggle(request, pk):
    try:
        obj = get_object_or_404(CustomerGrade, pk=pk, delete_time=0)
        obj.status = 1 - obj.status
        obj.save()
        
        status_text = '启用' if obj.status == 1 else '禁用'
        return JsonResponse({'code': 0, 'msg': f'客户等级已{status_text}'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'操作失败: {str(e)}'})


@login_required
def customer_grade_delete(request, pk):
    try:
        obj = get_object_or_404(CustomerGrade, pk=pk, delete_time=0)
        obj.delete_time = int(timezone.now().timestamp())
        obj.save()
        
        return JsonResponse({'code': 0, 'msg': '删除成功'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'删除失败: {str(e)}'})


@login_required
def customer_intent_list(request):
    return generic_list_view(
        request,
        CustomerIntent,
        'customer/customer_intent_list.html',
        search_fields=['name']
    )


@login_required
def customer_intent_form(request, pk=None):
    return generic_form_view(
        request,
        CustomerIntent,
        CustomerIntentForm,
        'customer/customer_intent_form.html',
        'basedata:customer_intent_list',
        pk
    )


@login_required
def customer_intent_list_data(request):
    try:
        search = request.GET.get('search', '')
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 20))
        
        queryset = CustomerIntent.objects.filter(delete_time=0)
        
        if search:
            queryset = queryset.filter(name__icontains=search)
        
        queryset = queryset.order_by('sort', 'id')
        
        paginator = Paginator(queryset, limit)
        page_obj = paginator.get_page(page)
        
        data = []
        for obj in page_obj:
            data.append({
                'id': obj.id,
                'name': obj.name,
                'sort': obj.sort,
                'status': obj.status,
                'status_display': '启用' if obj.status == 1 else '禁用',
                'create_time': obj.create_time.strftime('%Y-%m-%d %H:%M:%S') if obj.create_time else ''
            })
        
        return JsonResponse({
            'code': 0,
            'msg': '',
            'count': paginator.count,
            'data': data
        })
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'获取失败: {str(e)}'})


@login_required
def customer_intent_toggle(request, pk):
    try:
        obj = get_object_or_404(CustomerIntent, pk=pk, delete_time=0)
        obj.status = 1 - obj.status
        obj.save()
        
        status_text = '启用' if obj.status == 1 else '禁用'
        return JsonResponse({'code': 0, 'msg': f'客户意向已{status_text}'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'操作失败: {str(e)}'})


@login_required
def customer_intent_delete(request, pk):
    try:
        obj = get_object_or_404(CustomerIntent, pk=pk, delete_time=0)
        obj.delete_time = int(timezone.now().timestamp())
        obj.save()
        
        return JsonResponse({'code': 0, 'msg': '删除成功'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'删除失败: {str(e)}'})


@login_required
def customer_intent_update_sort(request, pk):
    try:
        obj = get_object_or_404(CustomerIntent, pk=pk, delete_time=0)
        new_sort = int(request.POST.get('sort', 0))
        
        obj.sort = new_sort
        obj.save()
        
        return JsonResponse({'code': 0, 'msg': '排序更新成功'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'排序更新失败: {str(e)}'})


@login_required
def customer_intent_batch_update_sort(request):
    try:
        data = json.loads(request.body)
        intents = data.get('intents', [])
        
        for item in intents:
            intent_id = item.get('id')
            new_sort = item.get('sort')
            
            if intent_id and new_sort is not None:
                CustomerIntent.objects.filter(id=intent_id, delete_time=0).update(sort=new_sort)
        
        return JsonResponse({'code': 0, 'msg': '批量排序更新成功'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'批量排序更新失败: {str(e)}'})


@login_required
def customer_field_list(request):
    return generic_list_view(
        request,
        CustomerField,
        'customer/customer_field_list.html',
        search_fields=['name', 'field_name']
    )


@login_required
def customer_field_form(request, pk=None):
    return generic_form_view(
        request,
        CustomerField,
        CustomerFieldForm,
        'customer/customer_field_form.html',
        'basedata:customer_field_list',
        pk
    )


@login_required
def customer_field_list_data(request):
    try:
        search = request.GET.get('search', '')
        page = int(request.GET.get('page', 1))
        limit = int(request.GET.get('limit', 20))
        
        queryset = CustomerField.objects.filter(delete_time=0)
        
        if search:
            queryset = queryset.filter(Q(name__icontains=search) | Q(field_name__icontains=search))
        
        queryset = queryset.order_by('sort', 'id')
        
        paginator = Paginator(queryset, limit)
        page_obj = paginator.get_page(page)
        
        data = []
        for obj in page_obj:
            data.append({
                'id': obj.id,
                'name': obj.name,
                'field_name': obj.field_name,
                'field_type': obj.field_type,
                'field_type_display': obj.get_field_type_display(),
                'is_required': obj.is_required,
                'is_unique': obj.is_unique,
                'is_list_display': obj.is_list_display,
                'sort': obj.sort,
                'status': obj.status,
                'status_display': '启用' if obj.status == 1 else '禁用',
                'create_time': obj.create_time.strftime('%Y-%m-%d %H:%M:%S') if obj.create_time else ''
            })
        
        return JsonResponse({
            'code': 0,
            'msg': '',
            'count': paginator.count,
            'data': data
        })
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'获取失败: {str(e)}'})


@login_required
def customer_field_toggle(request, pk):
    try:
        obj = get_object_or_404(CustomerField, pk=pk, delete_time=0)
        obj.status = 1 - obj.status
        obj.save()
        
        status_text = '启用' if obj.status == 1 else '禁用'
        return JsonResponse({'code': 0, 'msg': f'客户字段已{status_text}'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'操作失败: {str(e)}'})


@login_required
def customer_field_delete(request, pk):
    try:
        obj = get_object_or_404(CustomerField, pk=pk, delete_time=0)
        obj.delete_time = int(timezone.now().timestamp())
        obj.save()
        
        return JsonResponse({'code': 0, 'msg': '删除成功'})
    except Exception as e:
        return JsonResponse({'code': 1, 'msg': f'删除失败: {str(e)}'})


@login_required
def follow_field_list(request):
    return generic_list_view(
        request,
        FollowField,
        'customer/follow_field_list.html',
        search_fields=['name', 'field_name']
    )


@login_required
def follow_field_form(request, pk=None):
    return generic_form_view(
        request,
        FollowField,
        FollowFieldForm,
        'customer/follow_field_form.html',
        'customer:follow_field_list',
        pk
    )


@login_required
def order_field_list(request):
    return generic_list_view(
        request,
        OrderField,
        'customer/order_field_list.html',
        search_fields=['name', 'field_name']
    )


@login_required
def order_field_form(request, pk=None):
    return generic_form_view(
        request,
        OrderField,
        OrderFieldForm,
        'customer/order_field_form.html',
        'customer:order_field_list',
        pk
    )



